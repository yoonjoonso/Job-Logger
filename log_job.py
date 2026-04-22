from __future__ import annotations

import argparse
import json
import os
import random
import re
import difflib
import shutil
import sqlite3
import subprocess
import sys
import tempfile
import urllib.error
import urllib.request
from datetime import datetime
from functools import lru_cache
from html import unescape
from pathlib import Path

from policy_config import load_dynamic_archetype_catalog, normalize_search_text

# Ensure UTF-8 I/O on Windows (cp1252 default can't encode many Unicode chars in job postings)
if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
        sys.stdin.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass


VALID_AI = ("claude", "codex", "none")
JOB_POLICY_DB_PATH = Path(__file__).resolve().parent / "data" / "job-log.db"


def _load_archetype_runtime_catalog() -> dict[str, object]:
    return load_dynamic_archetype_catalog(Path(__file__).resolve().parent, JOB_POLICY_DB_PATH)


_ARCHETYPE_CATALOG = _load_archetype_runtime_catalog()
ARCHETYPE_ORDER = tuple(_ARCHETYPE_CATALOG["keys"])
DB_ARCHETYPE_NAMES = dict(_ARCHETYPE_CATALOG["labels"])
ARCHETYPE_FIT_MODIFIER: dict[str, float] = dict(_ARCHETYPE_CATALOG["fit_modifiers"])


def archetype_fit_modifier(archetype: str) -> float:
    """Return the fit-tier score modifier for a given archetype key."""
    return ARCHETYPE_FIT_MODIFIER.get(archetype, -0.5)


def get_repo_root() -> Path:
    override = os.environ.get("JOB_LOG_REPO_ROOT")
    if override:
        return Path(override).resolve()
    return Path(__file__).resolve().parent


def get_job_policy_db_path() -> Path:
    override = os.environ.get("JOB_LOG_POLICY_DB_PATH")
    if override:
        return Path(override).resolve()
    return get_repo_root() / "data" / "job-log.db"


def empty_job_policy() -> dict[str, object]:
    return {
        "source": "safety-fallback",
        "archetype_rules": {key: [] for key in ARCHETYPE_ORDER},
        "keyword_candidates": [],
        "keyword_blocklist": {},
    }


@lru_cache(maxsize=1)
def load_job_policy() -> dict[str, object]:
    db_path = get_job_policy_db_path()
    if not db_path.is_file():
        return empty_job_policy()

    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        tables = {
            row["name"]
            for row in connection.execute("SELECT name FROM sqlite_master WHERE type IN ('table', 'view')").fetchall()
        }
        if not {
            "job_archetype_rules",
            "job_keyword_candidates",
            "job_keyword_stopwords",
        }.issubset(tables):
            return empty_job_policy()

        archetype_rows = connection.execute(
            """
            SELECT archetype_key, keyword_text, keyword_normalized, weight, sort_priority
            FROM job_archetype_rules
            WHERE approved = 1
            ORDER BY archetype_key, sort_priority, keyword_text
            """
        ).fetchall()
        candidate_rows = connection.execute(
            """
            SELECT keyword_text, keyword_normalized, sort_priority
            FROM job_keyword_candidates
            WHERE approved = 1
            ORDER BY sort_priority, keyword_text
            """
        ).fetchall()
        stopword_rows = connection.execute(
            """
            SELECT term_text, term_normalized, kind
            FROM job_keyword_stopwords
            WHERE approved = 1
            ORDER BY kind, term_text
            """
        ).fetchall()

    archetype_rules: dict[str, list[dict[str, object]]] = {key: [] for key in ARCHETYPE_ORDER}
    for row in archetype_rows:
        archetype_key = normalize_text(row["archetype_key"])
        if archetype_key not in archetype_rules:
            continue
        archetype_rules[archetype_key].append(
            {
                "keyword": row["keyword_text"],
                "normalized": row["keyword_normalized"] or normalize_text(row["keyword_text"]),
                "weight": int(row["weight"] or 0),
                "sort_priority": int(row["sort_priority"] or 0),
            }
        )

    keyword_candidates = [
        {
            "keyword": row["keyword_text"],
            "normalized": row["keyword_normalized"] or normalize_text(row["keyword_text"]),
            "sort_priority": int(row["sort_priority"] or 0),
        }
        for row in candidate_rows
    ]
    keyword_blocklist = {
        row["term_normalized"] or normalize_text(row["term_text"]): {
            "term": row["term_text"],
            "kind": row["kind"] or "stopword",
        }
        for row in stopword_rows
    }
    return {
        "source": "db",
        "archetype_rules": archetype_rules,
        "keyword_candidates": keyword_candidates,
        "keyword_blocklist": keyword_blocklist,
    }


class AmbiguousExistingSelection(Exception):
    def __init__(self, rows: list[sqlite3.Row]) -> None:
        self.rows = rows
        super().__init__("Multiple matching roles found.")


def ai_extraction_schema() -> dict[str, object]:
    nullable_string = {"anyOf": [{"type": "string"}, {"type": "null"}]}
    nullable_number = {"anyOf": [{"type": "number"}, {"type": "null"}]}
    string_array = {
        "type": "array",
        "items": {"type": "string"},
    }
    return {
        "type": "object",
        "additionalProperties": False,
        "required": [
            "company",
            "role",
            "location",
            "work_model",
            "compensation_text",
            "salary_min",
            "salary_max",
            "salary_currency",
            "salary_period",
            "archetype",
            "jd_id",
            "keywords",
            "skills",
            "capabilities",
            "qualifications",
        ],
        "properties": {
            "company": nullable_string,
            "role": nullable_string,
            "location": nullable_string,
            "work_model": {"anyOf": [{"enum": ["remote", "hybrid", "onsite"]}, {"type": "null"}]},
            "compensation_text": nullable_string,
            "salary_min": nullable_number,
            "salary_max": nullable_number,
            "salary_currency": {"anyOf": [{"enum": ["USD", "CAD", "EUR", "GBP"]}, {"type": "null"}]},
            "salary_period": {"anyOf": [{"enum": ["yearly", "monthly", "hourly"]}, {"type": "null"}]},
            "archetype": {"enum": list(ARCHETYPE_ORDER)},
            "jd_id": nullable_string,
            "keywords": {
                "type": "array",
                "minItems": 3,
                "maxItems": 12,
                "items": {"type": "string"},
            },
            "skills": string_array,
            "capabilities": string_array,
            "qualifications": string_array,
        },
    }


_SUBSCORE_WEIGHTS: dict[str, float] = {
    "cv_match":  0.50,
    "role_fit":  0.28,
    "comp":      0.13,
    "work_pref": 0.09,
}


def compute_score_from_subscores(
    cv_match: float | None,
    role_fit: float | None,
    comp: float | None,
    work_pref: float | None,
    red_flag_penalty: float = 0.0,
) -> float | None:
    """Weighted average of available sub-scores minus red_flag_penalty.

    Weights are normalized to the sub-scores that are present so a role with
    only cv_match + role_fit still produces a meaningful score. Returns None
    when no sub-scores are populated (caller should fall back to keyword heuristic).
    red_flag_penalty is never modified here — it is always agent-set.
    """
    present = {
        k: v for k, v in [
            ("cv_match", cv_match),
            ("role_fit", role_fit),
            ("comp", comp),
            ("work_pref", work_pref),
        ] if v is not None
    }
    if not present:
        return None
    total_weight = sum(_SUBSCORE_WEIGHTS[k] for k in present)
    raw = sum(v * _SUBSCORE_WEIGHTS[k] / total_weight for k, v in present.items())
    return normalize_role_score(raw - red_flag_penalty)


def rescore_all_roles(repo_root: Path) -> int:
    """Recalculate scores for all DB roles using stored salary/work_model/requirements."""
    db_path = repo_root / "data" / "job-log.db"
    policy = load_job_policy()
    updated = 0
    skipped = 0

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        _ensure_score_pinned_column(conn)
        roles = conn.execute("SELECT * FROM roles ORDER BY num").fetchall()
        print(f"  Rescoring {len(roles)} roles...")
        print(_hr())

        for row in roles:
            notes = row["notes"] or ""
            # Skip roles with a pinned score (set via --score) or legacy pin in notes
            if row["score_pinned"] or extract_score_from_notes(notes) is not None:
                skipped += 1
                continue

            # Use JD text for classification when available — much more reliable than company+role
            jd_text = (row["jd_text"] or "") if "jd_text" in row.keys() else ""
            basis = jd_text if jd_text else " ".join(filter(None, [row["company"], row["role"], row["status"], notes]))
            classification = classify_job_description(basis, policy, include_details=True)
            matched_weight = sum(m["weight"] for m in classification["analysis"]["matched_rules"])
            stored_arch = (row["archetype"] or "") if "archetype" in row.keys() else ""
            # Prefer stored archetype; for NULL-arch entries without JD text, default to neutral (0.0)
            # to avoid unreliable guesses from sparse company+role text
            if stored_arch in ARCHETYPE_FIT_MODIFIER:
                arch = stored_arch
            elif jd_text:
                arch = classification["archetype"]
            else:
                arch = ""  # neutral — no reliable signal

            # Prefer sub-score weighted formula when agent has populated them
            penalty = float(_row_get(row, "red_flag_penalty") or 0)
            subscore_result = compute_score_from_subscores(
                _row_get(row, "cv_match"),
                _row_get(row, "role_fit"),
                _row_get(row, "comp"),
                _row_get(row, "work_pref"),
                penalty,
            )
            if subscore_result is not None:
                new_score = subscore_result
            else:
                # Fall back to keyword heuristic when no sub-scores present
                req_count = conn.execute(
                    "SELECT COUNT(*) FROM role_requirements WHERE role_id = ?", (row["num"],)
                ).fetchone()[0]
                keyword_count = req_count if req_count > 0 else len(extract_keywords(basis, policy))

                seniority_penalty = 0.5 if re.search(
                    r"\b(lead|principal|staff|manager|director)\b", row["role"] or "", re.IGNORECASE
                ) else 0.0

                fit_bonus = 0.0
                salary_max = float(row["salary_max"] or 0)
                salary_min = float(row["salary_min"] or 0)
                if salary_max > 0 and salary_max <= 80000:
                    fit_bonus -= 2.0
                elif salary_max > 0 and salary_max < 90000:
                    fit_bonus -= 1.0
                elif salary_min >= 100000:
                    fit_bonus += 0.5
                wm = row["work_model"] or ""
                if wm == "remote":
                    fit_bonus += 0.4
                elif wm == "hybrid":
                    fit_bonus += 0.1
                elif wm == "onsite":
                    fit_bonus -= 0.3

                raw = 2.5 + min(3.5, matched_weight / 5) + min(2.0, keyword_count / 5) + fit_bonus + archetype_fit_modifier(arch) - seniority_penalty
                new_score = normalize_role_score(raw)
            old_score = float(row["score"]) if row["score"] is not None else None

            conn.execute("UPDATE roles SET score = ? WHERE num = ?", (new_score, row["num"]))
            delta = ""
            if old_score is not None:
                diff = new_score - old_score
                if diff != 0:
                    delta = f"  ({old_score:g} -> {new_score:g}  Δ{diff:+.1f})"
                else:
                    delta = f"  ({new_score:g} unchanged)"
            print(f"  #{row['num']:<4} {row['company']:<28} {row['role'][:30]:<30}{delta}")
            updated += 1

        conn.commit()

    print(_hr())
    print(f"  Updated {updated} role scores. Skipped {skipped} (pinned score in notes).")
    return 0


def generate_report(repo_root: Path, month: str | None = None) -> int:
    """Generate an Excel summary of all logged roles, saved to output/{YYYY-MM}/job-log-report.xlsx."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        print("ERROR: openpyxl is required for --report. Install with: python -m pip install openpyxl")
        return 1

    db_path = repo_root / "data" / "job-log.db"
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        if month:
            rows = conn.execute(
                "SELECT * FROM roles WHERE date LIKE ? ORDER BY date DESC, num DESC",
                (f"{month}%",),
            ).fetchall()
        else:
            rows = conn.execute("SELECT * FROM roles ORDER BY date DESC, num DESC").fetchall()

    if not rows:
        label = f" for {month}" if month else ""
        print(f"  No roles found{label}.")
        return 0

    # Determine output folder
    if month:
        out_folder = repo_root / "output" / month
    else:
        out_folder = repo_root / "output" / datetime.now().strftime("%Y-%m")
    out_folder.mkdir(parents=True, exist_ok=True)
    out_path = out_folder / "job-log-report.xlsx"

    COLUMNS = [
        ("ID",           "num"),
        ("Date",         "date"),
        ("Company",      "company"),
        ("Role",         "role"),
        ("Status",       "status"),
        ("Score",        "score"),
        ("Archetype",    "archetype"),
        ("Work Model",   "work_model"),
        ("Location",     "location_text"),
        ("Compensation", "compensation_text"),
        ("Salary Min",   "salary_min"),
        ("Salary Max",   "salary_max"),
        ("Found Via",    "found_via"),
        ("Apply Method", "apply_method"),
        ("Source",       "source"),
        ("URL",          "url"),
        ("Notes",        "notes"),
    ]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roles"

    HEADER_FILL = PatternFill("solid", fgColor="1F3864")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
    ALT_FILL    = PatternFill("solid", fgColor="EEF2F7")

    # Header row
    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill   = HEADER_FILL
        cell.font   = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # Data rows
    keys = [r[0] for r in conn.execute("PRAGMA table_info(roles)").fetchall()] if False else None
    row_keys = [f for (_, f) in COLUMNS]
    for r_idx, row in enumerate(rows, start=2):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, field in enumerate(row_keys, start=1):
            try:
                val = row[field]
            except IndexError:
                val = None
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="top", wrap_text=(field == "notes"))

    # Column widths
    COL_WIDTHS = {
        "ID": 6, "Date": 12, "Company": 28, "Role": 32, "Status": 12, "Score": 8,
        "Archetype": 16, "Work Model": 12, "Location": 18, "Compensation": 20,
        "Salary Min": 11, "Salary Max": 11, "Found Via": 12, "Apply Method": 14,
        "Source": 10, "URL": 40, "Notes": 40,
    }
    for col_idx, (label, _) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = COL_WIDTHS.get(label, 16)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(out_path)
    label = f" ({month})" if month else f" ({len(rows)} total)"
    print(f"  Report saved: {out_path}{label}")
    return 0


def _resolve_inventory_name(
    connection: sqlite3.Connection,
    kind: str,
    raw: str,
) -> tuple[str, str] | None:
    """Resolve a skill/capability/qualification name or alias to (canonical_normalized, display_name).

    Tries in order: exact normalized match, exact name match (case-insensitive), alias match.
    Returns None if nothing is found.
    """
    needle = normalize_text(raw)
    if kind == "skill":
        canon_table, name_col, norm_col, alias_table, alias_fk = (
            "skills_mine", "skill_name", "skill_normalized", "skill_aliases", "skill_normalized"
        )
    elif kind == "capability":
        canon_table, name_col, norm_col, alias_table, alias_fk = (
            "capabilities_mine", "capability_name", "capability_normalized", "capability_aliases", "capability_normalized"
        )
    else:
        canon_table, name_col, norm_col, alias_table, alias_fk = (
            "qualifications_mine", "qualification_name", "qualification_normalized", "qualification_aliases", "qualification_normalized"
        )

    # Exact normalized match
    row = connection.execute(
        f"SELECT {norm_col} AS norm, {name_col} AS display FROM {canon_table} WHERE {norm_col} = ?", (needle,)
    ).fetchone()
    if row:
        return row["norm"], row["display"]

    # Case-insensitive name match
    row = connection.execute(
        f"SELECT {norm_col} AS norm, {name_col} AS display FROM {canon_table} WHERE lower({name_col}) = lower(?)", (raw,)
    ).fetchone()
    if row:
        return row["norm"], row["display"]

    # Alias match
    row = connection.execute(
        f"""SELECT a.{alias_fk} AS norm, c.{name_col} AS display
            FROM {alias_table} a JOIN {canon_table} c ON a.{alias_fk} = c.{norm_col}
            WHERE a.alias_normalized = ? OR lower(a.alias_name) = lower(?)""",
        (needle, raw),
    ).fetchone()
    if row:
        return row["norm"], row["display"]

    return None


def update_inventory(repo_root: Path, args: argparse.Namespace) -> int:
    db_path = repo_root / "data" / "job-log.db"
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row

        if args.set_skill:
            result = _resolve_inventory_name(conn, "skill", args.set_skill)
            if not result:
                print(f"  Error: no skill found matching '{args.set_skill}'.")
                return 1
            norm, display = result
            if args.level:
                conn.execute("UPDATE skills_mine SET level = ? WHERE skill_normalized = ?", (args.level, norm))
            if args.rename:
                conn.execute("UPDATE skills_mine SET skill_name = ? WHERE skill_normalized = ?", (args.rename, norm))
            conn.commit()
            row = conn.execute("SELECT skill_name, skill_normalized, level, resume_visibility FROM skills_mine WHERE skill_normalized = ?", (norm,)).fetchone()
            print(f"  Skill updated: {display}")
            print(f"    name:        {row['skill_name']}")
            print(f"    normalized:  {row['skill_normalized']}")
            print(f"    level:       {row['level']}")
            print(f"    visibility:  {row['resume_visibility']}")

        elif args.set_capability:
            result = _resolve_inventory_name(conn, "capability", args.set_capability)
            if not result:
                print(f"  Error: no capability found matching '{args.set_capability}'.")
                return 1
            norm, display = result
            if args.level:
                conn.execute("UPDATE capabilities_mine SET level = ? WHERE capability_normalized = ?", (args.level, norm))
            if args.rename:
                conn.execute("UPDATE capabilities_mine SET capability_name = ? WHERE capability_normalized = ?", (args.rename, norm))
            conn.commit()
            row = conn.execute("SELECT capability_name, capability_normalized, level FROM capabilities_mine WHERE capability_normalized = ?", (norm,)).fetchone()
            print(f"  Capability updated: {display}")
            print(f"    name:        {row['capability_name']}")
            print(f"    normalized:  {row['capability_normalized']}")
            print(f"    level:       {row['level']}")

        elif args.set_qualification:
            result = _resolve_inventory_name(conn, "qualification", args.set_qualification)
            if not result:
                print(f"  Error: no qualification found matching '{args.set_qualification}'.")
                return 1
            norm, display = result
            if args.met:
                conn.execute("UPDATE qualifications_mine SET met = ? WHERE qualification_normalized = ?", (args.met, norm))
            if args.rename:
                conn.execute("UPDATE qualifications_mine SET qualification_name = ? WHERE qualification_normalized = ?", (args.rename, norm))
            conn.commit()
            row = conn.execute("SELECT qualification_name, qualification_normalized, met, notes FROM qualifications_mine WHERE qualification_normalized = ?", (norm,)).fetchone()
            print(f"  Qualification updated: {display}")
            print(f"    name:        {row['qualification_name']}")
            print(f"    normalized:  {row['qualification_normalized']}")
            print(f"    met:         {row['met']}")
            if row["notes"]:
                print(f"    notes:       {row['notes']}")

    return 0


def main() -> int:
    # Ensure stdout/stderr can handle full Unicode on Windows (cp1252 default can't)
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    if hasattr(sys.stderr, "reconfigure"):
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    args = parse_args()
    repo_root = get_repo_root()
    job_policy = load_job_policy()
    if getattr(args, "rescore_all", False):
        return rescore_all_roles(repo_root)
    if getattr(args, "report", False):
        return generate_report(repo_root, month=getattr(args, "month", None))
    if getattr(args, "inventory_update", False):
        return update_inventory(repo_root, args)
    if args.update:
        if args.ai:
            row = refresh_existing_role_with_ai(repo_root, args, job_policy)
        else:
            _step(f"Updating role #{args.id}...")
            row = update_existing_role(
                repo_root,
                args.id,
                status=args.status,
                notes=args.notes,
                score=args.score,
                company=args.set_company,
                role=args.set_role,
                location_text=args.set_location,
                work_model=args.set_work_model,
                compensation_text=args.set_compensation_text,
                found_via=getattr(args, "set_found_via", None),
                apply_method=getattr(args, "set_apply_method", None),
            )
        print_existing_lookup(row, heading="Updated Role", verbose=args.verbose, repo_root=repo_root, show_jd=args.jd)
        if not (args.resume or args.coverletter or args.email):
            return 0
        # Fall through to artifact generation using the updated row as the record
        record = build_record_from_row(row, job_policy, archetype_override=getattr(args, "archetype", None))
        ai_provider = None
        source_mode = "existing"
    else:
        source_mode = get_source_mode(args)

        if args.minimal:
            record = build_minimal_record(args)
            ai_provider = None
        elif source_mode in {"url", "paste", "jdfile"}:
            record = build_new_intake_record(args, job_policy)
            ai_provider = args.ai if args.ai != "none" else None
        else:
            _step("Looking up existing role...")
            record = build_existing_record(args, repo_root, job_policy)
            ai_provider = None
            # Implicit status update: --id N --status X without --update
            if args.status and not args.resume and not args.coverletter and not args.email:
                _step(f"Updating status for #{record['db_row']['num']} -> {args.status}...")
                row = update_existing_role(repo_root, record["db_row"]["num"], status=args.status, notes=None, score=None)
                print_existing_lookup(row, heading="Updated Role", verbose=args.verbose, repo_root=repo_root, show_jd=args.jd)
                return 0
            # --set-archetype: persist archetype to DB
            if getattr(args, "set_archetype", None):
                _step(f"Setting archetype for #{record['db_row']['num']} -> {args.set_archetype}...")
                row = update_existing_role(repo_root, record["db_row"]["num"], status=None, notes=None, score=None, archetype=args.set_archetype)
                print_existing_lookup(row, heading="Updated Role", verbose=args.verbose, repo_root=repo_root, show_jd=args.jd)
                if not (args.resume or args.coverletter or args.email):
                    return 0
                record["archetype"] = args.set_archetype
            print_existing_lookup(record["db_row"], heading="Matched Role", verbose=args.verbose, repo_root=repo_root, show_jd=args.jd)
            if not (args.resume or args.coverletter or args.email):
                return 0

    company = record["company"]
    role = record["role"]
    location = record["location"]
    archetype = record["archetype"]
    keywords = record["keywords"]
    output_dir = repo_root / "output" / datetime.now().strftime("%Y-%m") / f"{slugify(company)}-{slugify(role)}"

    db_row = record.get("db_row")
    metadata = {
        "ai_provider": ai_provider,
        "source": record["source"],
        "company": company,
        "role": role,
        "date": record.get("date"),
        "location": location,
        "archetype": archetype,
        "keywords": keywords,
        "output_dir": str(output_dir),
        "compensation": record.get("compensation_text") or (record.get("compensation") or {}).get("text"),
        "work_model": record.get("work_model"),
        "jd_id": record.get("jd_id"),
        "resume": {
            "requested": args.resume,
            "status": "not_requested" if not args.resume else "pending",
            "pdf": None,
        },
        "coverletter": {
            "requested": args.coverletter,
            "status": "not_requested" if not args.coverletter else "pending",
            "path": None,
        },
        "email": {
            "requested": args.email,
            "status": "not_requested" if not args.email else "pending",
            "path": None,
        },
        "db": {
            "enabled": (source_mode in {"url", "paste", "jdfile"} or getattr(args, "minimal", False)) and not args.no_db,
            "status": "skipped" if args.no_db or source_mode in {"random", "existing"} else "pending",
            "role_id": db_row["num"] if isinstance(db_row, sqlite3.Row) else None,
            "role_status": db_row["status"] if isinstance(db_row, sqlite3.Row) else None,
            "role_score": db_row["score"] if isinstance(db_row, sqlite3.Row) else None,
        },
    }
    if record.get("analysis"):
        metadata["analysis"] = record["analysis"]
    if record.get("jd_text"):
        metadata["analysis_input"] = {
            "keyword_basis": record["jd_text"],
        }
    if record.get("analysis_input"):
        metadata["analysis_input"] = record["analysis_input"]
    if isinstance(db_row, sqlite3.Row):
        metadata["db"]["row"] = {key: db_row[key] for key in db_row.keys()}

    if args.dry_run:
        print_summary(metadata, dry_run=True, verbose=False)
        return 0

    output_dir.mkdir(parents=True, exist_ok=True)

    jd_text = record.get("jd_text")
    if jd_text:
        (output_dir / "jd.md").write_text(str(jd_text).rstrip() + "\n", encoding="utf-8")

    if metadata["db"]["enabled"]:
        if args.status:
            record["initial_status"] = args.status

        # Duplicate detection for DB-backed intake paths
        db_path = repo_root / "data" / "job-log.db"
        if db_path.is_file():
            with sqlite3.connect(db_path) as _dup_conn:
                _dup_conn.row_factory = sqlite3.Row
                _existing = _dup_conn.execute(
                    "SELECT num, company, role, status, score FROM roles WHERE company = ? AND role = ?",
                    (company, role),
                ).fetchone()
            if _existing:
                print(_hr())
                print(f"  Duplicate detected: #{_existing['num']}  {_existing['company']}  |  {_existing['role']}")
                print(f"  Status: {_existing['status'] or 'n/a'}  |  Score: {format_role_score(_existing['score'])}/10")
                print(_hr())
                duplicate_action = getattr(args, "on_duplicate", "prompt")
                if duplicate_action == "prompt":
                    print("  [1] Create new entry   [2] Update existing   [3] Cancel")
                    _choice = input("  Choice [1/2/3]: ").strip()
                elif duplicate_action == "create":
                    print("  Duplicate policy: create new entry.")
                    _choice = "1"
                elif duplicate_action == "update":
                    print(f"  Duplicate policy: update existing role #{_existing['num']}.")
                    _choice = "2"
                else:
                    print("  Duplicate policy: skip.")
                    _choice = "3"
                if _choice == "3":
                    print("  Cancelled.")
                    return 0
                elif _choice == "2":
                    # Update the existing role with new data and fall through to artifact generation
                    _step(f"Updating existing role #{_existing['num']}...")
                    _upd_row = update_existing_role(
                        repo_root,
                        int(_existing["num"]),
                        status=record.get("initial_status"),
                        notes=None,
                        score=None,
                    )
                    metadata["db"]["status"] = "logged"
                    metadata["db"]["role_id"] = int(_existing["num"])
                    if args.verbose:
                        metadata["db"]["full_row"] = fetch_full_db_row(repo_root, int(_existing["num"]))
                    # Use existing row's data for artifact generation
                    record["num"] = int(_existing["num"])
                    record["archetype"] = record.get("archetype") or extract_archetype_from_row(_upd_row)
                    print_existing_lookup(_upd_row, heading="Updated Role", verbose=False)
                    # Skip the log_to_db path
                    goto_artifact_gen = True
                else:
                    # choice "1": create new, fall through to log_to_db
                    goto_artifact_gen = False
            else:
                goto_artifact_gen = False
        else:
            goto_artifact_gen = False

        if not goto_artifact_gen:
            import sys as _sys
            if record.get("archetype") == "general" and _sys.stdin.isatty():
                arch_list = ", ".join(a for a in ARCHETYPE_ORDER if a != "general")
                print(f"\n  Archetype resolved to 'general'. Options: {arch_list}")
                override = input("  Enter archetype to override, or press Enter to keep 'general': ").strip().lower()
                if override and override in ARCHETYPE_ORDER:
                    record["archetype"] = override
                elif override:
                    print(f"  '{override}' not recognized — keeping 'general'.")
            # Collect found_via / apply_method / confirm status
            _db_path_for_sources = repo_root / "data" / "job-log.db"
            if source_mode in {"url", "paste", "jdfile"} or getattr(args, "minimal", False):
                _found_via = prompt_found_via(_db_path_for_sources, prefill=getattr(args, "found_via", None))
                if _found_via:
                    record["found_via"] = _found_via
                _explicit_how = getattr(args, "apply_method", None)
                _current_status = record.get("initial_status")  # set if --status was given
                if _explicit_how and not _current_status:
                    # --how implies applied
                    record["initial_status"] = "Applied"
                    record["apply_method"] = _explicit_how
                elif not _current_status and sys.stdin.isatty():
                    _status_ans = input("\n  Applied or just evaluating? [A/e]: ").strip().lower()
                    print()
                    if _status_ans.startswith("e"):
                        record["initial_status"] = "Evaluated"
                    else:
                        record["initial_status"] = "Applied"
                        _apply_method = prompt_apply_method(prefill=_explicit_how)
                        if _apply_method:
                            record["apply_method"] = _apply_method
                elif _current_status == "Applied":
                    _apply_method = prompt_apply_method(prefill=_explicit_how)
                    if _apply_method:
                        record["apply_method"] = _apply_method
            _step("Writing to database...")
            role_id, new_candidates = log_to_db(repo_root, record, ai_provider or "", minimal=bool(record.get("minimal")))
            metadata["db"]["status"] = "logged"
            metadata["db"]["role_id"] = role_id
            if args.verbose:
                metadata["db"]["full_row"] = fetch_full_db_row(repo_root, role_id)
            prompt_new_inventory_items(new_candidates, repo_root / "data" / "job-log.db")

    if args.resume:
        _step("Generating resume PDF...")
        resume_pdf = run_resume_generation(repo_root, record, output_dir)
        metadata["resume"]["status"] = "generated"
        metadata["resume"]["pdf"] = str(resume_pdf)
        resume_metadata = load_resume_metadata(output_dir)
        if resume_metadata:
            metadata["resume"]["metadata_path"] = str(output_dir / "resume-metadata.json")
            metadata["resume"]["selection"] = {
                "profile": (resume_metadata.get("explainability") or {}).get("profile_selection"),
                "layout": (resume_metadata.get("explainability") or {}).get("layout_selection"),
                "education": (resume_metadata.get("explainability") or {}).get("education_selection"),
                "certifications": (resume_metadata.get("explainability") or {}).get("certification_selection"),
                "experience": (resume_metadata.get("explainability") or {}).get("experience_selection"),
                "derived_features": (resume_metadata.get("explainability") or {}).get("derived_features"),
            }
        if not args.no_open:
            _step("Opening PDF...")
            open_file(resume_pdf)

    if args.coverletter:
        _step("Generating cover letter...")
        cl_out = run_coverletter_generation(repo_root, record, output_dir)
        metadata["coverletter"]["status"] = "generated"
        metadata["coverletter"]["docx"]   = str(cl_out["docx"])
        metadata["coverletter"]["pdf"]    = str(cl_out["pdf"]) if cl_out.get("pdf") else None
        if not args.no_open:
            open_file(cl_out.get("pdf") or cl_out["docx"])

    if args.email:
        _step("Generating email...")
        email_path = run_email_generation(repo_root, record, output_dir)
        metadata["email"]["status"] = "generated"
        metadata["email"]["path"]   = str(email_path)
        if not args.no_open:
            open_file(email_path)

    (output_dir / "metadata.json").write_text(json.dumps(metadata, indent=2), encoding="utf-8")
    print_summary(metadata, dry_run=False, verbose=getattr(args, "verbose", False))
    return 0


_LEVEL_NAMES = ["none", "exposure", "basic", "intermediate", "advanced", "expert"]


def _parse_level(val: str) -> str:
    if val.isdigit():
        n = int(val)
        if 0 <= n <= 5:
            return _LEVEL_NAMES[n]
        raise argparse.ArgumentTypeError(f"numeric level must be 0–5, got {n}")
    if val in _LEVEL_NAMES:
        return val
    raise argparse.ArgumentTypeError(f"invalid level '{val}'; choose a name or 0–5")


class _HelpFormatter(argparse.HelpFormatter):
    """HelpFormatter that inserts a horizontal rule before each argument group."""

    def format_help(self) -> str:
        text = super().format_help()
        lines = text.split("\n")
        out: list[str] = []
        sep = "─" * min(self._width, 90)
        first_heading = True
        for line in lines:
            is_heading = bool(line and not line.startswith(" ") and line.rstrip().endswith(":"))
            if is_heading:
                if not first_heading:
                    while out and out[-1] == "":
                        out.pop()
                    out.append("")
                    out.append(sep)
                first_heading = False
            out.append(line)
        return "\n".join(out) + "\n"


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Job intake, lookup, and artifact generation pipeline.",
        formatter_class=lambda prog: _HelpFormatter(prog, max_help_position=32, width=95),
    )

    # ── Intake ─────────────────────────────────────────────────────────────────
    g_intake = parser.add_argument_group("intake  (log a new role)")
    g_intake.add_argument("--url",      metavar="URL",  help="Fetch and log a role from a job posting URL.")
    g_intake.add_argument("--paste",    action="store_true", help="Paste a JD manually; end input with a line containing only END.")
    g_intake.add_argument("--jdfile",   metavar="FILE", help="Load a JD from a local text or markdown file.")
    g_intake.add_argument("--minimal",  action="store_true", help="Log a bare open application with no JD (--company required; --role/--archetype optional).")
    g_intake.add_argument("--company",  metavar="NAME", help="Company name (required for --minimal; override for other intake modes).")
    g_intake.add_argument("--role",     metavar="TITLE", help="Role title override for intake or selector for company-based lookup.")
    g_intake.add_argument("--ai",       choices=VALID_AI, metavar="{claude,codex,none}", help="AI provider for JD analysis (default: none).")
    g_intake.add_argument("--via",      dest="found_via",    metavar="SOURCE", help="Where you found the role (e.g. linkedin, company, referral).")
    g_intake.add_argument("--how",      dest="apply_method", metavar="METHOD", choices=["company", "linkedin", "email", "other"], help="How you applied: company, linkedin, email, or other.")
    g_intake.add_argument("--on-duplicate", choices=["prompt", "create", "update", "skip"], default="prompt",
                          help="When a duplicate company+role exists during intake: prompt (default), create, update, or skip.")
    g_intake.add_argument("--no-db",    action="store_true", help="Skip writing to the database.")

    # ── Lookup ─────────────────────────────────────────────────────────────────
    g_lookup = parser.add_argument_group("lookup  (find an existing role)")
    g_lookup.add_argument("--id",       type=int, metavar="N", help="Look up or update a logged role by numeric ID.")
    g_lookup.add_argument("--random",   action="store_true",   help="Pick a random existing DB role.")

    # ── Artifacts ──────────────────────────────────────────────────────────────
    g_art = parser.add_argument_group("artifacts  (generate output files)")
    g_art.add_argument("--resume",      action="store_true", help="Generate a tailored resume (DOCX + PDF).")
    g_art.add_argument("--coverletter", action="store_true", help="Generate a cover letter (DOCX + PDF).")
    g_art.add_argument("--email",       action="store_true", help="Generate a short outreach email body (TXT).")
    g_art.add_argument("--archetype",   choices=list(ARCHETYPE_ORDER), metavar="ARCHETYPE",
                       help="Override archetype for artifact generation. Choices: " + ", ".join(ARCHETYPE_ORDER) + ".")
    g_art.add_argument("--no-open",     action="store_true", help="Skip auto-opening generated files after creation.")
    g_art.add_argument("--dry-run",     action="store_true", help="Show planned actions without writing files.")

    # ── Update ─────────────────────────────────────────────────────────────────
    g_update = parser.add_argument_group("update  (modify a logged role; most require --id)")
    g_update.add_argument("--status",               metavar="STATUS", help="Set role status.")
    g_update.add_argument("--notes",                metavar="TEXT",   help="Replace role notes.")
    g_update.add_argument("--score",                type=float, metavar="N", help="Manually set role score.")
    g_update.add_argument("--set-company",          metavar="NAME",   help="Update company name.")
    g_update.add_argument("--set-role",             metavar="TITLE",  help="Update role title.")
    g_update.add_argument("--set-location",         metavar="TEXT",   help="Update location text.")
    g_update.add_argument("--set-work-model",       choices=["remote", "hybrid", "onsite"], metavar="{remote,hybrid,onsite}", help="Update work model.")
    g_update.add_argument("--set-compensation-text", metavar="TEXT",  help="Update compensation text.")
    g_update.add_argument("--set-via",  dest="set_found_via",    metavar="SOURCE", help="Update found_via.")
    g_update.add_argument("--set-how",  dest="set_apply_method", metavar="METHOD", choices=["company", "linkedin", "email", "other"], help="Update apply_method.")
    g_update.add_argument("--set-archetype", choices=list(ARCHETYPE_ORDER), metavar="ARCHETYPE", help="Permanently store archetype for a role.")

    # ── Inventory ──────────────────────────────────────────────────────────────
    g_inv = parser.add_argument_group("inventory  (manage skills / capabilities / qualifications)")
    g_inv.add_argument("--set-skill",         metavar="NAME", help="Set a skill level (use with --level).")
    g_inv.add_argument("--set-capability",    metavar="NAME", help="Set a capability level (use with --level).")
    g_inv.add_argument("--set-qualification", metavar="NAME", help="Set a qualification met status (use with --met).")
    g_inv.add_argument("--level",  type=_parse_level,
                       choices=["none", "exposure", "basic", "intermediate", "advanced", "expert"],
                       metavar="{0-5 | none,exposure,basic,intermediate,advanced,expert}",
                       help="Level for --set-skill / --set-capability (0=none … 5=expert).")
    g_inv.add_argument("--met",    choices=["yes", "partial", "no"], metavar="{yes,partial,no}", help="Met status for --set-qualification.")
    g_inv.add_argument("--rename", metavar="NEW_NAME", help="Rename a skill/capability/qualification display name.")
    g_inv.add_argument("--rescore-all", action="store_true", help="Recalculate scores for all DB roles.")

    # ── Output ─────────────────────────────────────────────────────────────────
    g_out = parser.add_argument_group("output")
    g_out.add_argument("--verbose", "-v", action="store_true", help="Print full DB row and extracted requirements.")
    g_out.add_argument("--jd",            action="store_true", help="Print the stored JD text (use with --id).")
    g_out.add_argument("--report",        action="store_true", help="Generate an Excel summary of all logged roles.")
    g_out.add_argument("--month",         metavar="YYYY-MM",   help="Filter --report to a specific month (default: all).")

    args = parser.parse_args()

    if getattr(args, "rescore_all", False):
        return args

    if getattr(args, "report", False):
        return args

    has_inventory_update = bool(args.set_skill or args.set_capability or args.set_qualification)
    if has_inventory_update:
        if args.set_skill and not args.level and not args.rename:
            parser.error("--set-skill requires --level or --rename.")
        if args.set_capability and not args.level and not args.rename:
            parser.error("--set-capability requires --level or --rename.")
        if args.set_qualification and not args.met and not args.rename:
            parser.error("--set-qualification requires --met or --rename.")
        args.inventory_update = True
        return args
    args.inventory_update = False

    has_intake_source = bool(args.url or args.paste or args.jdfile or args.minimal)
    has_update_fields = any([
        args.notes is not None,
        args.score is not None,
        args.set_company is not None,
        args.set_role is not None,
        args.set_location is not None,
        args.set_work_model is not None,
        args.set_compensation_text is not None,
        getattr(args, "set_found_via", None) is not None,
        getattr(args, "set_apply_method", None) is not None,
    ])
    is_ai_refresh = args.id is not None and has_intake_source and bool(args.ai)
    is_update = args.id is not None and (has_update_fields or is_ai_refresh)
    is_lookup = args.id is not None or args.company or args.random

    if has_update_fields and args.id is None:
        parser.error("--set-*, --notes, and --score require --id.")
    if is_ai_refresh and has_update_fields:
        parser.error("Cannot mix manual field updates with AI source refresh.")
    if has_intake_source and not args.minimal and not args.ai and not is_ai_refresh:
        parser.error("New JD intake modes require --ai {claude|codex|none}.")
    if args.minimal and not args.company:
        parser.error("--minimal requires --company.")

    if args.ai and not has_intake_source:
        parser.error("--ai is only valid with --url, --paste, or --jdfile.")
    if not is_update and not is_lookup and not has_intake_source:
        parser.error("Provide a source (--url/--paste/--jdfile/--minimal/--random) or a lookup target (--id/--company).")
    if args.role and not args.company and not has_intake_source:
        parser.error("--role requires --company for lookup, or a new intake source.")
    if args.no_db and has_update_fields:
        parser.error("--no-db is not valid with field update flags.")

    CANONICAL_STATUSES = {"Evaluated", "Applied", "Responded", "Interview", "Offer", "Rejected", "Discarded", "SKIP"}
    if args.status:
        normalized = args.status.upper() if args.status.lower() == "skip" else args.status.capitalize()
        if normalized not in CANONICAL_STATUSES:
            parser.error(f"--status must be one of: {', '.join(sorted(CANONICAL_STATUSES))}")
        args.status = normalized

    # Synthesise update flag for main() to consume
    args.update = is_update

    return args


def get_source_mode(args: argparse.Namespace) -> str:
    if args.url:
        return "url"
    if args.paste:
        return "paste"
    if args.jdfile:
        return "jdfile"
    if args.random:
        return "random"
    return "existing"


def extract_archetype_from_row(row: sqlite3.Row) -> str:
    """Derive archetype from a DB row using deterministic extraction."""
    text = " ".join(filter(None, [
        row["role"] if "role" in row.keys() else "",
        row["notes"] if "notes" in row.keys() else "",
    ]))
    return classify_job_description(text, load_job_policy(), include_details=False)


def build_record_from_row(row: sqlite3.Row, job_policy: dict[str, object], archetype_override: str | None = None) -> dict[str, object]:
    """Build a minimal record dict from an existing DB row for artifact generation."""
    keys = row.keys()
    location = (row["location_text"] if "location_text" in keys else None) or ""
    work_model = row["work_model"] if "work_model" in keys else ""
    comp_text = row["compensation_text"] if "compensation_text" in keys else ""
    # Use stored archetype > CLI override > deterministic fallback from role title
    stored_archetype = (row["archetype"] if "archetype" in keys else None) or None
    if archetype_override and archetype_override in ARCHETYPE_ORDER:
        archetype = archetype_override
    elif stored_archetype and stored_archetype in ARCHETYPE_ORDER:
        archetype = stored_archetype
    else:
        text_basis = " ".join(filter(None, [row["role"], location, comp_text, row["notes"]]))
        archetype = deterministic_extract(text_basis, job_policy)["archetype"]
    text_basis = " ".join(filter(None, [row["role"], location, comp_text, row["notes"]]))
    extracted = deterministic_extract(text_basis, job_policy)
    return {
        "source": {"type": "existing", "value": f"role#{row['num']}"},
        "jd_text": None,
        "company": row["company"],
        "role": row["role"],
        "date": row["date"],
        "location": location,
        "location_text": location,
        "work_model": work_model or "",
        "compensation_text": comp_text or "",
        "archetype": archetype,
        "keywords": extracted["keywords"],
        "requirements": extracted.get("requirements") or {"skills": [], "capabilities": [], "qualifications": []},
        "db_row": row,
        "analysis": extracted.get("analysis") or {},
    }


def build_minimal_record(args: argparse.Namespace) -> dict[str, object]:
    """Build a bare record for an open application with no JD."""
    archetype = getattr(args, "archetype", None) or "general"
    return {
        "source": {"type": "minimal", "value": "manual"},
        "jd_text": "",
        "company": args.company.strip(),
        "role": (args.role or "Open Application").strip(),
        "location": "",
        "work_model": "",
        "compensation": {},
        "archetype": archetype,
        "keywords": [],
        "requirements": {"skills": [], "capabilities": [], "qualifications": []},
        "jd_id": None,
        "analysis": {},
        "minimal": True,
    }


def build_new_intake_record(args: argparse.Namespace, job_policy: dict[str, object]) -> dict[str, object]:
    if args.url:
        _step(f"Fetching job description from URL...")
        source = {"type": "url", "value": args.url}
        jd_text = read_from_url(args.url)
    elif args.paste:
        source = {"type": "paste", "value": "interactive"}
        jd_text = read_from_paste()
    else:
        _step(f"Reading job description from file...")
        path = Path(args.jdfile).resolve()
        source = {"type": "jdfile", "value": str(path)}
        jd_text = read_from_file(path)

    jd_text = clean_jd_text(jd_text)

    if args.ai and args.ai != "none":
        _step(f"Calling {args.ai} to extract metadata (this may take a moment)...")
        extracted = extract_with_ai(args.ai, jd_text, job_policy, get_repo_root())
    else:
        _step("Extracting metadata (deterministic)...")
        extracted = deterministic_extract(jd_text, job_policy)
    company = args.company or extracted.get("company")
    role = args.role or extracted.get("role")
    if not company or not str(company).strip():
        raise SystemExit("Could not determine company. Pass --company \"Name\" explicitly.")
    if not role or not str(role).strip():
        raise SystemExit("Could not determine role. Pass --role \"Title\" explicitly.")
    return {
        "source": source,
        "jd_text": jd_text,
        "company": str(company).strip(),
        "role": str(role).strip(),
        "location": extracted["location"] or "",
        "work_model": extracted.get("work_model") or "",
        "compensation": extracted.get("compensation") or {},
        "archetype": extracted["archetype"],
        "keywords": extracted["keywords"],
        "requirements": extracted.get("requirements") or {"skills": [], "capabilities": [], "qualifications": []},
        "jd_id": extracted.get("jd_id") or None,
        "analysis": extracted.get("analysis") or {},
    }


def build_existing_record(args: argparse.Namespace, repo_root: Path, job_policy: dict[str, object]) -> dict[str, object]:
    db_path = repo_root / "data" / "job-log.db"
    if not db_path.is_file():
        raise SystemExit(f"DB not found: {db_path}")

    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        row = ensure_row_has_score(connection, resolve_source_row(connection, args))
        keywords = extract_existing_keywords(connection, row, job_policy)
        role_requirements = connection.execute(
            """
            SELECT source, requirement_name, kind, priority, matched_entity_type, matched_normalized, match_method, confidence, notes
            FROM role_requirements
            WHERE role_id = ?
            ORDER BY source, requirement_name
            """,
            (row["num"],),
        ).fetchall()

    text_basis = " ".join(
        filter(
            None,
            [
                row["role"],
                row["location_text"] if "location_text" in row.keys() else "",
                row["compensation_text"] if "compensation_text" in row.keys() else "",
                row["report"] if "report" in row.keys() else "",
                row["notes"],
            ],
        )
    )
    extracted = deterministic_extract(text_basis, job_policy)
    # Prefer stored archetype > CLI override > deterministic fallback
    stored_archetype = (row["archetype"] if "archetype" in row.keys() else None) or None
    archetype_override = getattr(args, "archetype", None)
    if archetype_override and archetype_override in ARCHETYPE_ORDER:
        resolved_archetype = archetype_override
    elif stored_archetype and stored_archetype in ARCHETYPE_ORDER:
        resolved_archetype = stored_archetype
    else:
        resolved_archetype = extracted["archetype"]
    role_row = {key: row[key] for key in row.keys()}
    requirement_dicts = [dict(req) for req in role_requirements]
    return {
        "source": {"type": source_label(args), "value": f"role#{row['num']}"},
        "jd_text": None,
        "company": row["company"],
        "role": row["role"],
        "date": row["date"],
        "location": row["location_text"] if "location_text" in row.keys() else extract_location(row["notes"] or "") or "",
        "work_model": row["work_model"] if "work_model" in row.keys() else "",
        "archetype": resolved_archetype,
        "keywords": keywords,
        "requirements": {"skills": [], "capabilities": [], "qualifications": []},
        "db_row": row,
        "analysis": extracted.get("analysis") or {},
        "analysis_input": {
            "stored_role": role_row,
            "keyword_basis": text_basis,
            "role_requirements": requirement_dicts,
            "matched_requirements": [req for req in requirement_dicts if req.get("matched_entity_type") and req["matched_entity_type"] != "none"],
            "derived_requirements": [req for req in requirement_dicts if req.get("source") == "jd_keywords" or req.get("match_method") == "new_candidate"],
        },
    }


def resolve_source_row(connection: sqlite3.Connection, args: argparse.Namespace) -> sqlite3.Row:
    if args.random:
        rows = connection.execute(
            "SELECT * FROM roles ORDER BY num"
        ).fetchall()
        if not rows:
            raise SystemExit("No logged roles found in the DB for --random.")
        return random.choice(rows)

    if args.id is not None:
        row = connection.execute(
            "SELECT * FROM roles WHERE num = ?",
            (args.id,),
        ).fetchone()
        if row is None:
            raise SystemExit(f'No DB role found for id "{args.id}".')
        return row

    return resolve_existing_row(connection, args.company, args.role)


def resolve_existing_row(connection: sqlite3.Connection, company: str, role: str | None) -> sqlite3.Row:
    rows = connection.execute(
        "SELECT * FROM roles WHERE lower(company) LIKE ? ORDER BY num DESC",
        (f"%{company.lower()}%",),
    ).fetchall()

    if not rows:
        raise SystemExit(f'No DB role found for company "{company}".')

    if role:
        ranked = sorted(
            rows,
            key=lambda row: (
                -match_score(row["company"] or "", company),
                -match_score(row["role"] or "", role),
                -(row["num"] or 0),
            ),
        )
        if match_score(ranked[0]["role"] or "", role) == 0:
            raise SystemExit(f'No DB role found for company "{company}" with role matching "{role}".')
        return ranked[0]

    exact_company_rows = [row for row in rows if normalize_text(row["company"] or "") == normalize_text(company)]
    if len(exact_company_rows) == 1:
        return exact_company_rows[0]
    if len(exact_company_rows) > 1:
        raise AmbiguousExistingSelection(sorted(exact_company_rows, key=lambda row: -(row["num"] or 0)))
    if len(rows) == 1:
        return rows[0]
    raise AmbiguousExistingSelection(rows)


def source_label(args: argparse.Namespace) -> str:
    if args.random:
        return "random"
    return "existing"


def load_jd_source(args: argparse.Namespace) -> tuple[dict[str, str], str]:
    if args.url:
        return {"type": "url", "value": args.url}, read_from_url(args.url)
    if args.paste:
        return {"type": "paste", "value": "interactive"}, read_from_paste()
    if args.jdfile:
        path = Path(args.jdfile).resolve()
        return {"type": "jdfile", "value": str(path)}, read_from_file(path)
    raise SystemExit("No JD source provided.")


def find_conflicting_role(connection: sqlite3.Connection, company: str, role: str, exclude_id: int) -> sqlite3.Row | None:
    return connection.execute(
        "SELECT num, company, role, status, score FROM roles WHERE company = ? AND role = ? AND num <> ?",
        (company, role, exclude_id),
    ).fetchone()


def update_existing_role(
    repo_root: Path,
    role_id: int,
    status: str | None,
    notes: str | None,
    score: float | None,
    company: str | None = None,
    role: str | None = None,
    location_text: str | None = None,
    work_model: str | None = None,
    compensation_text: str | None = None,
    archetype: str | None = None,
    found_via: str | None = None,
    apply_method: str | None = None,
) -> sqlite3.Row:
    db_path = repo_root / "data" / "job-log.db"
    if not db_path.is_file():
        raise SystemExit(f"DB not found: {db_path}")

    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        _ensure_score_pinned_column(connection)
        existing = connection.execute(
            "SELECT * FROM roles WHERE num = ?",
            (role_id,),
        ).fetchone()
        if existing is None:
            raise SystemExit(f'No DB role found for id "{role_id}".')

        next_company = existing["company"] if company is None else clean_label(company)
        next_role = existing["role"] if role is None else clean_label(role)
        if not next_company:
            raise SystemExit("Updated company cannot be empty.")
        if not next_role:
            raise SystemExit("Updated role cannot be empty.")
        conflict = find_conflicting_role(connection, next_company, next_role, role_id)
        if conflict is not None:
            raise SystemExit(
                f'Update would conflict with existing role #{conflict["num"]}: {conflict["company"]} | {conflict["role"]}.'
            )

        next_status = existing["status"] if status is None else status
        next_notes = existing["notes"] if notes is None else notes
        next_score = existing["score"] if score is None else normalize_role_score(score)
        next_score_pinned = 1 if score is not None else (existing["score_pinned"] if "score_pinned" in existing.keys() else 0)
        next_location = existing["location_text"] if location_text is None else (clean_label(location_text) or None)
        next_work_model = existing["work_model"] if work_model is None else work_model
        next_compensation = existing["compensation_text"] if compensation_text is None else (clean_label(compensation_text) or None)
        _ensure_archetype_column(connection)
        _ensure_found_via_columns(connection)
        existing_archetype = (existing["archetype"] if "archetype" in existing.keys() else None) or None
        next_archetype = existing_archetype if archetype is None else archetype
        existing_cols = {row[1] for row in connection.execute("PRAGMA table_info(roles)").fetchall()}
        next_found_via = (existing["found_via"] if "found_via" in existing_cols else None) if found_via is None else found_via
        next_apply_method = (existing["apply_method"] if "apply_method" in existing_cols else None) if apply_method is None else apply_method
        connection.execute(
            """
            UPDATE roles
            SET company = ?, role = ?, status = ?, notes = ?, score = ?, score_pinned = ?,
                location_text = ?, work_model = ?, compensation_text = ?, last_updated_date = ?,
                archetype = ?, found_via = ?, apply_method = ?
            WHERE num = ?
            """,
            (next_company, next_role, next_status, next_notes, next_score, next_score_pinned, next_location, next_work_model, next_compensation, today_iso(), next_archetype, next_found_via, next_apply_method, role_id),
        )
        connection.commit()

        updated = ensure_row_has_score(
            connection,
            connection.execute(
                "SELECT * FROM roles WHERE num = ?",
                (role_id,),
            ).fetchone(),
        )
        if updated is None:
            raise SystemExit(f'Updated role "{role_id}" could not be reloaded.')
        return updated


def refresh_existing_role_with_ai(repo_root: Path, args: argparse.Namespace, job_policy: dict[str, object]) -> sqlite3.Row:
    db_path = repo_root / "data" / "job-log.db"
    if not db_path.is_file():
        raise SystemExit(f"DB not found: {db_path}")

    _step("Fetching job description...")
    source, jd_text = load_jd_source(args)
    _step(f"Calling {args.ai} to refresh metadata (this may take a moment)...")
    extracted = extract_with_ai(args.ai, jd_text, job_policy, repo_root)
    company = extracted.get("company")
    role = extracted.get("role")
    if not company or not role:
        raise SystemExit("AI refresh did not return a valid company and role. Aborting without updating.")

    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        existing = connection.execute("SELECT * FROM roles WHERE num = ?", (args.id,)).fetchone()
        if existing is None:
            raise SystemExit(f'No DB role found for id "{args.id}".')

        conflict = find_conflicting_role(connection, company, role, args.id)
        if conflict is not None:
            raise SystemExit(
                f'AI refresh for #{args.id} would conflict with existing role #{conflict["num"]}: {conflict["company"]} | {conflict["role"]}.'
            )

        compensation = extracted.get("compensation") or {}
        next_score = normalize_role_score(args.score) if args.score is not None else derive_role_score_from_text(
            role,
            jd_text,
            compensation,
            extracted.get("work_model"),
            extracted.get("keywords", []),
            job_policy,
        )
        next_notes = existing["notes"] if args.notes is None else args.notes
        next_status = existing["status"] if args.status is None else args.status

        _step("Updating database...")
        connection.execute(
            """
            UPDATE roles
            SET company = ?, role = ?, location_text = ?, work_model = ?, compensation_text = ?,
                salary_min = ?, salary_max = ?, salary_currency = ?, salary_period = ?,
                score = ?, status = ?, notes = ?, source = ?, via = ?, last_updated_date = ?
            WHERE num = ?
            """,
            (
                company,
                role,
                extracted.get("location"),
                extracted.get("work_model"),
                compensation.get("text"),
                compensation.get("salary_min"),
                compensation.get("salary_max"),
                compensation.get("salary_currency"),
                compensation.get("salary_period"),
                next_score,
                next_status,
                next_notes,
                source.get("type"),
                source.get("value"),
                today_iso(),
                args.id,
            ),
        )
        write_role_requirements(connection, args.id, role, jd_text, extracted.get("keywords", []), extracted.get("requirements"))
        connection.commit()

        updated = connection.execute("SELECT * FROM roles WHERE num = ?", (args.id,)).fetchone()
        if updated is None:
            raise SystemExit(f'Updated role "{args.id}" could not be reloaded.')
        return updated


def ensure_row_has_score(connection: sqlite3.Connection, row: sqlite3.Row | None) -> sqlite3.Row | None:
    if row is None or row["score"] is not None:
        return row

    derived_score = derive_role_score(row)
    connection.execute("UPDATE roles SET score = ? WHERE num = ?", (derived_score, row["num"]))
    connection.commit()
    return connection.execute(
        "SELECT * FROM roles WHERE num = ?",
        (row["num"],),
    ).fetchone()


def derive_role_score(row: sqlite3.Row) -> float:
    notes = row["notes"] or ""
    pre_score = extract_score_from_notes(notes)
    if pre_score is not None:
        return pre_score

    basis = " ".join(filter(None, [row["company"], row["role"], row["status"], notes]))
    classification = classify_job_description(basis, load_job_policy(), include_details=True)
    stored_arch = (row["archetype"] or "") if "archetype" in row.keys() else ""
    archetype = stored_arch if stored_arch in ARCHETYPE_FIT_MODIFIER else classification["archetype"]
    matched_weight = sum(match["weight"] for match in classification["analysis"]["matched_rules"])
    keyword_count = len(extract_keywords(basis, load_job_policy()))
    seniority_penalty = 0.5 if re.search(r"\b(lead|principal|staff|manager|director)\b", row["role"] or "", re.IGNORECASE) else 0.0
    raw_score = 2.5 + min(3.5, matched_weight / 5) + min(2.0, keyword_count / 5) + archetype_fit_modifier(archetype) - seniority_penalty
    return normalize_role_score(raw_score)


def extract_score_from_notes(notes: str) -> float | None:
    match = re.search(r"\b(?:pre[- ]score|score)\s*[:=]?\s*(\d+(?:\.\d+)?)\b", notes, re.IGNORECASE)
    if not match:
        return None
    return normalize_role_score(float(match.group(1)))


def normalize_role_score(value: float) -> float:
    clamped = max(0.0, min(10.0, float(value)))
    return round(clamped * 2) / 2


def match_score(value: str, needle: str) -> int:
    left = normalize_text(value)
    right = normalize_text(needle)
    if not left or not right:
        return 0
    if left == right:
        return 100
    if right in left:
        return 70
    left_tokens = set(left.split())
    right_tokens = set(right.split())
    return len(left_tokens & right_tokens) * 10


def deterministic_extract(text: str, job_policy: dict[str, object] | None = None) -> dict[str, object]:
    policy = job_policy or load_job_policy()
    classification = classify_job_description(text, policy, include_details=True)
    keyword_result = extract_keywords(text, policy, include_details=True)
    compensation = extract_compensation(text)
    return {
        "company": extract_company(text),
        "role": extract_role(text),
        "location": extract_location(text),
        "work_model": extract_work_model(text),
        "compensation": compensation,
        "archetype": classification["archetype"],
        "keywords": keyword_result["keywords"],
        "requirements": {"skills": [], "capabilities": [], "qualifications": []},
        "analysis": {
            "policy_source": policy.get("source"),
            "archetype_classification": classification["analysis"],
            "keyword_extraction": keyword_result["analysis"],
        },
    }


def build_ai_extraction_prompt(text: str, job_policy: dict[str, object]) -> str:
    policy_payload = {
        "archetype_rules": job_policy.get("archetype_rules", {}),
        "keyword_candidates": job_policy.get("keyword_candidates", []),
    }
    archetype_list = ", ".join(ARCHETYPE_ORDER)
    return "\n".join(
        [
            "Extract structured metadata from this job description.",
            "Return JSON only. Do not include markdown fences or commentary.",
            "Do not use generic headings like JOB DESCRIPTION as the company.",
            "If the employer is not explicit, return null for company rather than guessing.",
            f"Use one of these archetypes only: {archetype_list}.",
            "The 'keywords' field MUST be a JSON array of 5-12 strings. Never return null for keywords.",
            "Choose concrete JD keywords: technologies, platforms, languages, domains, and delivery terms over generic filler.",
            "Prefer explicit named software, APIs, engines, file formats, and platforms from required/preferred/bonus sections over generic concepts or location words.",
            "If the JD names concrete tools like Maya, Houdini, Unreal, Unity, OpenGL, Vulkan, DirectX, GLSL, HLSL, USD, or CUDA, include the most relevant ones in keywords.",
            "Also classify extracted requirements into three JSON string arrays: 'skills', 'capabilities', and 'qualifications'.",
            "Use 'skills' for named tools, APIs, languages, engines, platforms, file formats, and concrete software.",
            "Use 'capabilities' for broader functional areas like rendering pipelines, developer tooling, multithreading, visualization, or technical leadership.",
            "Use 'qualifications' for constraints or gatekeeping requirements like years of experience, degree requirements, work authorization, location restrictions, clearance, certifications, or onsite/hybrid constraints.",
            "If a bucket has no good entries, return an empty JSON array for that bucket.",
            "If compensation is present, preserve the raw compensation_text and parse salary fields when explicit. Otherwise use nulls.",
            "If the JD contains an explicit job/requisition/posting ID number (e.g. 'Req ID: 12345', 'Job #ABC-001'), extract it as jd_id. Otherwise return null.",
            "Policy context:",
            json.dumps(policy_payload, indent=2),
            "Job description:",
            text,
        ]
    )


def extract_json_object(raw: str) -> dict[str, object]:
    text = str(raw or "").strip()
    if not text:
        raise SystemExit("AI extraction returned empty output.")
    try:
        parsed = json.loads(text)
    except json.JSONDecodeError:
        match = re.search(r"\{.*\}", text, re.DOTALL)
        if not match:
            raise SystemExit(f"AI extraction did not return valid JSON: {text[:300]}")
        try:
            parsed = json.loads(match.group(0))
        except json.JSONDecodeError as exc:
            raise SystemExit(f"AI extraction returned invalid JSON: {exc}") from exc
    if not isinstance(parsed, dict):
        raise SystemExit("AI extraction JSON must be an object.")
    return parsed


def normalize_ai_string_list(value: object, *, min_items: int = 0, max_items: int = 12, label: str = "items") -> list[str]:
    if not isinstance(value, list):
        raise SystemExit(f"AI extraction returned invalid {label}; expected a JSON array.")
    items = []
    seen = set()
    for item in value:
        text = clean_label(str(item or ""))
        normalized = normalize_text(text)
        if not text or not normalized or normalized in seen:
            continue
        items.append(text)
        seen.add(normalized)
        if len(items) >= max_items:
            break
    if len(items) < min_items:
        raise SystemExit(f"AI extraction returned too few usable {label}. Aborting without logging.")
    return items


def normalize_ai_keyword_list(value: object) -> list[str]:
    return normalize_ai_string_list(value, min_items=3, max_items=12, label="keywords")


EXPLICIT_JD_KEYWORD_PATTERNS: list[tuple[str, str]] = [
    (r"\bmaya\b", "Maya"),
    (r"\bhoudini\b", "Houdini"),
    (r"\bopenusd\b|\busd\b", "USD"),
    (r"\bopengl\b", "OpenGL"),
    (r"\bvulkan\b", "Vulkan"),
    (r"\bdirectx\b", "DirectX"),
    (r"\bglsl\b", "GLSL"),
    (r"\bhlsl\b", "HLSL"),
    (r"\bcuda\b", "CUDA"),
    (r"\bunreal(?: engine)?\b", "Unreal"),
    (r"\bunity\b", "Unity"),
    (r"\bsimd\b", "SIMD"),
    (r"\bmultithread(?:ed|ing)?\b", "multithreading"),
]

LOW_SIGNAL_AI_KEYWORDS = {
    "burbank",
    "california",
    "united",
    "states",
    "source",
    "paste",
    "codex",
}


def supplement_ai_keywords_from_jd(jd_text: str | None, keywords: list[str]) -> list[str]:
    if not jd_text:
        return keywords

    result = list(keywords)
    seen = {normalize_text(item) for item in result if normalize_text(item)}
    explicit_terms = []
    for pattern, label in EXPLICIT_JD_KEYWORD_PATTERNS:
        if re.search(pattern, jd_text, re.IGNORECASE):
            normalized = normalize_text(label)
            if normalized and normalized not in seen:
                explicit_terms.append(label)
                seen.add(normalized)

    if not explicit_terms:
        return result

    for label in explicit_terms:
        if len(result) < 12:
            result.append(label)
            continue
        replacement_index = next(
            (idx for idx in range(len(result) - 1, -1, -1) if normalize_text(result[idx]) in LOW_SIGNAL_AI_KEYWORDS),
            None,
        )
        if replacement_index is None:
            break
        result[replacement_index] = label

    deduped = []
    deduped_seen = set()
    for item in result:
        normalized = normalize_text(item)
        if not normalized or normalized in deduped_seen:
            continue
        deduped.append(item)
        deduped_seen.add(normalized)
        if len(deduped) >= 12:
            break
    return deduped


def normalize_ai_requirement_groups(payload: dict[str, object]) -> dict[str, list[str]]:
    return {
        "skills": normalize_ai_string_list(payload.get("skills") or [], max_items=12, label="skills"),
        "capabilities": normalize_ai_string_list(payload.get("capabilities") or [], max_items=12, label="capabilities"),
        "qualifications": normalize_ai_string_list(payload.get("qualifications") or [], max_items=12, label="qualifications"),
    }


def normalize_ai_extraction(
    payload: dict[str, object],
    jd_text: str | None = None,
    job_policy: dict[str, object] | None = None,
) -> dict[str, object]:
    archetype = normalize_text(payload.get("archetype"))
    # "none" means AI returned null — try deterministic classifier on JD text first
    if archetype in ("none", "") and jd_text:
        det = classify_job_description(jd_text, job_policy or load_job_policy(), include_details=False)
        if isinstance(det, str) and det in ARCHETYPE_ORDER:
            archetype = det
    if archetype not in ARCHETYPE_ORDER:
        raw = str(payload.get("archetype") or "").lower()
        # also try to guess from JD text via classifier
        if jd_text:
            det = classify_job_description(jd_text, job_policy or load_job_policy(), include_details=False)
            guessed = det if isinstance(det, str) and det in ARCHETYPE_ORDER else next((k for k in ARCHETYPE_ORDER if k in raw), None)
        else:
            guessed = next((k for k in ARCHETYPE_ORDER if k in raw), None)
        print(f'\nAI returned archetype "{payload.get("archetype")}". Pick a valid one:')
        for i, key in enumerate(ARCHETYPE_ORDER, 1):
            hint = " <-- guess" if key == guessed else ""
            print(f"  {i}. {key} ({DB_ARCHETYPE_NAMES.get(key, key)}){hint}")
        print("  n. new  (enter a custom archetype key)")
        print("  c. cancel")
        while True:
            try:
                choice = input("Archetype [1-{} / key / enter to accept guess / n / c]: ".format(len(ARCHETYPE_ORDER))).strip().lower()
            except (EOFError, KeyboardInterrupt):
                raise SystemExit("Cancelled.")
            if choice == "c":
                raise SystemExit("Cancelled.")
            if choice == "n":
                try:
                    new_key = input("New archetype key (lowercase, no spaces): ").strip().lower()
                except (EOFError, KeyboardInterrupt):
                    raise SystemExit("Cancelled.")
                new_key = re.sub(r"[^a-z0-9_-]", "", new_key)
                if new_key:
                    archetype = new_key
                    break
                print("  Empty key — try again.")
                continue
            if choice == "" and guessed:
                archetype = guessed
                break
            if choice in ARCHETYPE_ORDER:
                archetype = choice
                break
            if choice.isdigit() and 1 <= int(choice) <= len(ARCHETYPE_ORDER):
                archetype = ARCHETYPE_ORDER[int(choice) - 1]
                break
            print("  Invalid — try again.")

    work_model = normalize_text(payload.get("work_model"))
    if work_model in ("none", "null", "n/a"):
        work_model = ""
    if work_model == "on-site":
        work_model = "onsite"
    if work_model not in {"remote", "hybrid", "onsite", ""}:
        raise SystemExit(f'AI extraction returned invalid work_model "{payload.get("work_model")}".')

    salary_currency = str(payload.get("salary_currency")).strip().upper() if payload.get("salary_currency") is not None else None
    if salary_currency not in {None, "USD", "CAD", "EUR", "GBP"}:
        raise SystemExit(f'AI extraction returned invalid salary_currency "{payload.get("salary_currency")}".')

    salary_period = normalize_text(payload.get("salary_period"))
    if salary_period in ("none", "null", "n/a"):
        salary_period = ""
    if salary_period not in {"yearly", "monthly", "hourly", ""}:
        raise SystemExit(f'AI extraction returned invalid salary_period "{payload.get("salary_period")}".')

    def as_string(value: object) -> str | None:
        if value is None:
            return None
        text = clean_label(str(value))
        return text or None

    def as_number(value: object) -> float | None:
        if value is None or value == "":
            return None
        try:
            return float(value)
        except (TypeError, ValueError) as exc:
            raise SystemExit(f'AI extraction returned invalid numeric value "{value}".') from exc

    return {
        "company": as_string(payload.get("company")),
        "role": as_string(payload.get("role")),
        "location": as_string(payload.get("location")),
        "work_model": work_model or None,
        "compensation": {
            "text": as_string(payload.get("compensation_text")),
            "salary_min": as_number(payload.get("salary_min")),
            "salary_max": as_number(payload.get("salary_max")),
            "salary_currency": salary_currency,
            "salary_period": salary_period or None,
        },
        "archetype": archetype,
        "keywords": normalize_ai_keyword_list(payload.get("keywords")),
        "requirements": normalize_ai_requirement_groups(payload),
        "jd_id": as_string(payload.get("jd_id")),
    }


def run_claude_extraction(prompt: str) -> dict[str, object]:
    claude_exe = shutil.which("claude")
    if not claude_exe:
        raise SystemExit("Claude CLI is not installed or not available in PATH.")
    command = [
        claude_exe,
        "-p",
        "--permission-mode",
        "default",
        "--output-format",
        "json",
        "--json-schema",
        json.dumps(ai_extraction_schema()),
    ]
    try:
        completed = subprocess.run(command, input=prompt, text=True, encoding="utf-8", capture_output=True, check=True)
    except FileNotFoundError as exc:
        raise SystemExit(f"Claude CLI not found at '{claude_exe}'. Verify your PATH.") from exc
    except subprocess.CalledProcessError as exc:
        message = exc.stderr.strip() or exc.stdout.strip() or "Claude extraction failed"
        raise SystemExit(f"Claude extraction failed: {message}") from exc
    stdout = completed.stdout.strip()
    if not stdout:
        stderr = completed.stderr.strip()
        raise SystemExit(f"Claude extraction returned empty output.{' stderr: ' + stderr if stderr else ''}")
    envelope = extract_json_object(stdout)
    # --json-schema puts the validated payload in "structured_output"
    structured = envelope.get("structured_output")
    if isinstance(structured, dict):
        return structured
    # fallback: plain --output-format json puts text in "result"
    raw_result = envelope.get("result")
    if isinstance(raw_result, dict):
        return raw_result
    if isinstance(raw_result, str) and raw_result.strip():
        return extract_json_object(raw_result)
    raise SystemExit("Claude extraction returned no usable payload (structured_output and result both empty).")


def run_codex_extraction(prompt: str, repo_root: Path) -> dict[str, object]:
    schema = ai_extraction_schema()
    schema_path = None
    output_path = None
    command = None
    try:
        codex_tmp_dir = repo_root / "temp"
        codex_tmp_dir.mkdir(parents=True, exist_ok=True)
        with tempfile.NamedTemporaryFile(
            "w",
            suffix="-codex-schema.json",
            prefix="job-log-",
            dir=str(codex_tmp_dir),
            delete=False,
            encoding="utf-8",
        ) as schema_file:
            json.dump(schema, schema_file, indent=2)
            schema_path = Path(schema_file.name)
        with tempfile.NamedTemporaryFile(
            "w",
            suffix="-codex-output.json",
            prefix="job-log-",
            dir=str(codex_tmp_dir),
            delete=False,
            encoding="utf-8",
        ) as output_file:
            output_path = Path(output_file.name)

        command = build_codex_command(
            "exec",
            "--skip-git-repo-check",
            "--sandbox",
            "read-only",
            "--output-schema",
            str(schema_path),
            "--output-last-message",
            str(output_path),
            "-",
        )
        try:
            subprocess.run(
                command,
                input=prompt,
                text=True,
                encoding="utf-8",
                capture_output=True,
                check=True,
                cwd=str(repo_root),
            )
        except FileNotFoundError as exc:
            if command and Path(command[0]).exists():
                raise SystemExit(
                    f"Codex CLI was found at {command[0]!r}, but execution failed before startup: {exc}"
                ) from exc
            raise SystemExit("Codex CLI is not installed or not available in PATH.") from exc
        except subprocess.CalledProcessError as exc:
            message = exc.stderr.strip() or exc.stdout.strip() or "Codex extraction failed"
            raise SystemExit(f"Codex extraction failed: {message}") from exc

        return extract_json_object(output_path.read_text(encoding="utf-8"))
    finally:
        for path in (schema_path, output_path):
            if path and path.exists():
                try:
                    path.unlink()
                except OSError:
                    pass


def build_codex_command(*args: str) -> list[str]:
    codex_path = shutil.which("codex") or shutil.which("codex.cmd") or shutil.which("codex.exe") or shutil.which("codex.ps1")
    if not codex_path:
        raise FileNotFoundError("codex")

    suffix = Path(codex_path).suffix.lower()
    if suffix == ".ps1":
        pwsh_path = shutil.which("pwsh") or shutil.which("powershell")
        if not pwsh_path:
            raise FileNotFoundError("pwsh")
        return [pwsh_path, "-NoProfile", "-ExecutionPolicy", "Bypass", "-File", codex_path, *args]
    return [codex_path, *args]


def extract_with_ai(ai_provider: str | None, text: str, job_policy: dict[str, object], repo_root: Path) -> dict[str, object]:
    if ai_provider not in VALID_AI:
        raise SystemExit(f"Unsupported AI provider: {ai_provider}")
    prompt = build_ai_extraction_prompt(text, job_policy)
    last_error: str | None = None
    for attempt in range(1, 3):
        if attempt > 1:
            _step(f"Retrying AI extraction (attempt {attempt}/2, previous error: {last_error})...")
            retry_prompt = (
                prompt
                + f"\n\nPrevious attempt failed validation: {last_error}"
                + "\nFix the issue and return valid JSON."
            )
            current_prompt = retry_prompt
        else:
            current_prompt = prompt
        if ai_provider == "claude":
            payload = run_claude_extraction(current_prompt)
        else:
            payload = run_codex_extraction(current_prompt, repo_root)
        try:
            normalized = normalize_ai_extraction(payload, jd_text=text, job_policy=job_policy)
            normalized["keywords"] = supplement_ai_keywords_from_jd(text, normalized.get("keywords", []))
            normalized.setdefault("requirements", {"skills": [], "capabilities": [], "qualifications": []})
            return {
                **normalized,
                "analysis": {
                    "provider": ai_provider,
                    "policy_source": job_policy.get("source"),
                    "raw_extraction": payload,
                },
            }
        except SystemExit as exc:
            last_error = str(exc)
            if attempt >= 2:
                raise
    raise SystemExit(f"AI extraction failed after retries: {last_error}")


def read_from_url(url: str) -> str:
    if not re.match(r"^https?://", url, re.IGNORECASE):
        raise SystemExit("URL input must start with http:// or https://")
    try:
        with urllib.request.urlopen(url, timeout=20) as response:
            html = response.read().decode("utf-8", errors="replace")
    except urllib.error.URLError as exc:
        raise SystemExit(f"Failed to fetch URL: {exc}") from exc
    text = html_to_text(html).strip()
    if not text:
        raise SystemExit("Fetched URL but could not extract readable text.")
    return text


def read_from_paste() -> str:
    print("Paste the job description. End with a line containing only END.")
    lines: list[str] = []
    while True:
        try:
            line = input()
        except EOFError:
            break
        if line.strip() == "END":
            break
        lines.append(line)
    text = "\n".join(lines).strip()
    if not text:
        raise SystemExit("No job description provided.")
    return text


def read_from_file(path: Path) -> str:
    if not path.is_file():
        raise SystemExit(f"JD file does not exist: {path}")
    text = path.read_text(encoding="utf-8")
    if not text.strip():
        raise SystemExit(f"JD file is empty: {path}")
    return text


def clean_jd_text(text: str) -> str:
    """Deterministic cosmetic cleanup of raw JD text. Does not alter content."""
    import unicodedata
    import re as _re
    # Normalize unicode compatibility equivalents (ligatures, superscripts, NBSP variants, etc.)
    text = unicodedata.normalize("NFKC", text)
    # Normalize line endings
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    # Normalize by Unicode category:
    #   Pd (dash punctuation) — any hyphen/dash variant → hyphen-minus
    #   Pi/Pf (initial/final quote) — curly quotes → straight
    #   Po bullets (•, ·, ▪, etc.) → hyphen-minus
    _BULLET_CHARS = {"\u2022", "\u2023", "\u2043", "\u204c", "\u204d", "\u2219", "\u25aa", "\u25ab", "\u25b8", "\u25cf", "\u25e6", "\u00b7", "\u2027"}
    def _normalize_char(ch: str) -> str:
        cat = unicodedata.category(ch)
        if cat == "Pd":
            return "-"
        if cat in ("Pi", "Pf"):
            return "'" if unicodedata.name(ch, "").startswith("SINGLE") else '"'
        if ch in _BULLET_CHARS:
            return "-"
        return ch
    text = "".join(_normalize_char(ch) for ch in text)
    # Replace any whitespace that isn't a newline (NBSP, thin space, tab, etc.) with a plain space
    text = _re.sub(r"[^\S\n]", " ", text)
    # Strip non-printable control characters (keep newlines)
    text = "".join(ch for ch in text if ch == "\n" or not unicodedata.category(ch).startswith("C"))
    # Strip trailing whitespace from each line
    lines = [line.rstrip() for line in text.split("\n")]
    # Collapse runs of 3+ blank lines down to 2
    result: list[str] = []
    blank_run = 0
    for line in lines:
        if line == "":
            blank_run += 1
            if blank_run <= 2:
                result.append(line)
        else:
            blank_run = 0
            result.append(line)
    return "\n".join(result).strip()


def classify_job_description(text: str, job_policy: dict[str, object] | None = None, include_details: bool = False) -> str | dict[str, object]:
    policy = job_policy or load_job_policy()
    normalized = normalize_text(text)
    scored: list[dict[str, object]] = []
    for archetype in ARCHETYPE_ORDER:
        rules = policy.get("archetype_rules", {}).get(archetype, [])
        score = 1 if archetype == "general" else 0
        matched_rules = []
        for rule in rules:
            keyword_normalized = str(rule.get("normalized") or normalize_text(rule.get("keyword")))
            if keyword_normalized and keyword_normalized in normalized:
                weight = int(rule.get("weight") or 0)
                score += weight
                matched_rules.append(
                    {
                        "keyword": rule.get("keyword"),
                        "keyword_normalized": keyword_normalized,
                        "weight": weight,
                    }
                )
        scored.append({"archetype": archetype, "score": score, "matched_rules": matched_rules})
    scored.sort(key=lambda item: (-int(item["score"]), ARCHETYPE_ORDER.index(str(item["archetype"]))))
    winner = scored[0] if scored else {"archetype": "general", "score": 1, "matched_rules": []}
    if not include_details:
        return str(winner["archetype"])
    return {
        "archetype": winner["archetype"],
        "analysis": {
            "selected_archetype": winner["archetype"],
            "policy_source": policy.get("source"),
            "matched_rules": winner["matched_rules"],
            "scores": scored,
        },
    }


def extract_keywords(text: str, job_policy: dict[str, object] | None = None, include_details: bool = False) -> list[str] | dict[str, object]:
    policy = job_policy or load_job_policy()
    normalized = normalize_text(text)
    found: list[str] = []
    seen: set[str] = set()
    blocked = dict(policy.get("keyword_blocklist", {}))
    candidate_matches: list[dict[str, object]] = []
    fallback_matches: list[dict[str, object]] = []
    blocked_terms: list[dict[str, object]] = []

    for candidate in policy.get("keyword_candidates", []):
        token = str(candidate.get("normalized") or normalize_text(candidate.get("keyword")))
        if token and token in normalized and token not in seen:
            found.append(str(candidate.get("keyword")))
            seen.add(token)
            candidate_matches.append(
                {
                    "keyword": candidate.get("keyword"),
                    "keyword_normalized": token,
                    "sort_priority": int(candidate.get("sort_priority") or 0),
                }
            )
        if len(found) >= 12:
            break

    words = re.findall(r"[a-z][a-z0-9+#/-]{2,}", text.lower())
    for word in words:
        if len(found) >= 12:
            break
        if word in blocked:
            blocked_terms.append({"term": word, "kind": blocked[word].get("kind", "stopword")})
            continue
        if word in seen:
            continue
        found.append(word)
        seen.add(word)
        fallback_matches.append({"keyword": word, "keyword_normalized": word, "source": "fallback_token"})
    if not include_details:
        return found
    return {
        "keywords": found,
        "analysis": {
            "policy_source": policy.get("source"),
            "selected_keywords": found,
            "candidate_matches": candidate_matches,
            "fallback_matches": fallback_matches,
            "blocked_terms": blocked_terms,
        },
    }


def split_keyword_terms(value: str) -> list[str]:
    parts = re.split(r"[,;/|+]+", str(value or ""))
    output = []
    for part in parts:
        cleaned = clean_label(part)
        if cleaned:
            output.append(cleaned)
    return output


def strip_experience_qualifier(value: str) -> str:
    return re.sub(
        r"\s+",
        " ",
        re.sub(
            r"\b\d+(?:\+|[–-]\d+)?\s*yrs?(?:\s+req)?\b",
            " ",
            re.sub(r"^\d+(?:\+|[–-]\d+)?\s*yrs?(?:\s+req)?\s+", "", str(value or ""), flags=re.IGNORECASE),
            flags=re.IGNORECASE,
        ),
    ).strip()


def split_generic_composite(value: str) -> list[str]:
    placeholder = "__cplusplus__"
    protected = re.sub(r"c\+\+", placeholder, str(value or ""), flags=re.IGNORECASE)
    parts = re.split(r"[+/]", protected)
    output: list[str] = []
    for part in parts:
        restored = re.sub(placeholder, "C++", part, flags=re.IGNORECASE)
        cleaned = clean_label(restored)
        if cleaned:
            output.append(cleaned)
    return output


SLASH_PRESERVE_TERMS = {
    "html/css",
    "tcp/ip",
    "ci/cd",
    "ci/cd and automation",
    "slo/sli",
    "vr/xr",
    "vr/xr development",
    "ts/sci",
    "ts/sci clearance",
}


def expand_requirement_term(term: str) -> list[str]:
    cleaned = clean_label(strip_experience_qualifier(term))
    normalized = re.sub(r"\s+", " ", re.sub(r"[()]", " ", normalize_text(cleaned))).strip()
    if not cleaned:
        return []
    composite_map = {
        "ci/cd": ["CI/CD and Automation"],
        "ci/cd jenkins/gitlab": ["CI/CD and Automation", "Jenkins", "GitLab CI"],
        "ci/cd jenkins/gitlab/github": ["CI/CD and Automation", "Jenkins", "GitLab CI", "GitHub Actions"],
        "docker/k8s": ["Docker", "Kubernetes"],
        "docker/k8s/terraform": ["Docker", "Kubernetes", "Terraform"],
        "docker/ci-cd/agile/junit": ["Docker", "CI/CD and Automation", "Agile", "JUnit"],
        ".net/asp.net": [".NET", "ASP.NET"],
        "c#/.net asp.net": ["C#", ".NET", "ASP.NET"],
        "c#/.net asp.net rest": ["C#", ".NET", "ASP.NET", "REST APIs"],
        "c#/.net asp.net core": ["C#", ".NET", "ASP.NET Core"],
        "c#/.net asp.net mvc": ["C#", ".NET", "ASP.NET MVC"],
        "c#/.net asp.net mvc rest": ["C#", ".NET", "ASP.NET MVC", "REST APIs"],
        "c#/.net backend": ["C#", ".NET", "Backend Frameworks"],
        "c#/.net primary": ["C#", ".NET"],
        "k8s/docker": ["Kubernetes", "Docker"],
        "k8s/helm/docker": ["Kubernetes", "Helm", "Docker"],
        "k8s/openshift": ["Kubernetes", "OpenShift"],
        "rest/websockets": ["REST APIs", "WebSockets"],
        "rest/grpc": ["REST APIs", "gRPC"],
        "rest microservices": ["REST APIs", "Microservices"],
        "sql/nosql": ["SQL", "NoSQL"],
        "sql+nosql": ["SQL", "NoSQL"],
        "sql/nosql/redis": ["SQL", "NoSQL", "Redis"],
        "git/perforce": ["Git", "Perforce"],
        "maven/gradle": ["Maven", "Gradle"],
        "junit/mockito": ["JUnit", "Mockito"],
        "javascript/typescript": ["JavaScript", "TypeScript"],
        "react": ["React"],
        "react/typescript frontend": ["React", "TypeScript"],
        "react/typescript/next.js": ["React", "TypeScript"],
        "react/angular/vue spa": ["React"],
        "react/angular .net/node/rust": ["React", ".NET", "Node.js"],
        "react/ts/node/go backend": ["React", "TypeScript", "Node.js", "Go"],
        "electron react/node bonus": ["React", "Node.js"],
        "node.js/java/python": ["Node.js", "Java", "Python"],
        "node/java/python backend": ["Node.js", "Java", "Python"],
        "python/node/c#": ["Python", "Node.js", "C#"],
        "python/go": ["Python", "Go"],
        "unity/unreal": ["Unity", "Unreal Engine"],
        "c/c++": ["C", "C++"],
        "c/c++/c#": ["C", "C++", "C#"],
        "c/c++/c#/python": ["C", "C++", "C#", "Python"],
        "c++/c#": ["C++", "C#"],
        "c++/java": ["C++", "Java"],
        "c#/c++/python": ["C#", "C++", "Python"],
        "python/c#": ["Python", "C#"],
        "python/c++": ["Python", "C++"],
        "python/c#/c++": ["Python", "C#", "C++"],
        "c++/python/js": ["C++", "Python", "JavaScript"],
        "python/go automation": ["Python", "Go"],
        "python/go primary": ["Python", "Go"],
        "python/go scripting": ["Python", "Go"],
        "python/go/java": ["Python", "Go", "Java"],
        "java/c++/python": ["Java", "C++", "Python"],
        "java/kotlin": ["Java", "Kotlin"],
        "java/docker/ci-cd/agile": ["Java", "Docker", "CI/CD and Automation", "Agile"],
        "java/j2ee/linux/microservices/devsecops/rmf": ["Java", "Linux", "Microservices"],
        "java/kotlin/microservices/distributed systems": ["Java", "Kotlin", "Microservices", "Distributed Systems"],
        "java/maven/gradle/git": ["Java", "Maven", "Gradle", "Git"],
        "html/css/js": ["HTML/CSS", "JavaScript"],
        "spring boot": ["Spring Boot"],
        "graphql": ["GraphQL"],
        "cloud aws/gcp/azure": ["Cloud Infrastructure", "AWS", "GCP", "Azure"],
        "cloud aws/azure": ["Cloud Infrastructure", "AWS", "Azure"],
        "aws/azure cloud": ["Cloud Infrastructure", "AWS", "Azure"],
        "aws/gcp": ["AWS", "GCP"],
        "aws/gcp/azure": ["AWS", "GCP", "Azure"],
        "aws/azure/gcp eks/ec2/s3": ["Cloud Infrastructure", "AWS", "Azure", "GCP"],
        "aws/k8s/docker/vmware": ["AWS", "Kubernetes", "Docker"],
        "iac terraform/ansible": ["Cloud Infrastructure", "Terraform", "Ansible"],
        "iac terraform/cloudformation": ["Cloud Infrastructure", "Terraform", "CloudFormation"],
        "iac terraform/packer/ansible": ["Cloud Infrastructure", "Terraform", "Packer", "Ansible"],
        "gitops argocd/flux": ["GitOps", "ArgoCD", "Flux"],
        "slo/sli": ["SLO/SLI"],
        "secret clearance": ["Security Clearance"],
        "secret clearance req": ["Security Clearance"],
        "ts/sci req": ["TS/SCI Clearance"],
        "ts/sci clearance": ["TS/SCI Clearance"],
        "slos/slis/error budgets": ["SLO/SLI"],
        "db": ["Database Systems"],
        "oop": ["OOP"],
        "gpu/cpu optimization": ["GPU Optimization", "CPU Optimization"],
        "anti-cheat/anti-malware": ["Anti-Cheat Systems", "Anti-Malware"],
        "graphics apis": ["Graphics APIs"],
        "graphics apis dx12/vulkan": ["Graphics APIs"],
        "graphics apis directx/opengl/metal": ["Graphics APIs"],
        "directx/vulkan": ["Graphics APIs"],
        "directx/vulkan/opengl": ["Graphics APIs"],
        "build systems": ["Build Systems"],
        "client/server": ["Client-Server Architecture"],
        "dev workflows": ["Development Workflows"],
        "large codebases": ["Large Codebase Maintenance"],
        "latency mitigation": ["Latency Mitigation"],
        "linear algebra": ["3D Math"],
        "low-level systems": ["Low-Level Systems"],
        "mentoring": ["Technical Mentoring"],
        "networking": ["Networking"],
        "rendering systems": ["Rendering Systems"],
        "shaders": ["Shaders"],
        "scalable distributed systems": ["Distributed Systems"],
        "gpu architecture": ["GPU Architecture"],
        "genai tools copilot": ["GenAI Tools"],
        "agile": ["Agile"],
        "perforce/git": ["Perforce", "Git"],
        "rest/json": ["REST APIs", "JSON"],
        "rest/grpc apis": ["REST APIs", "gRPC"],
        "rest/microservices": ["REST APIs", "Microservices"],
        "k8s/cloud": ["Kubernetes", "Cloud Infrastructure"],
        "k8s/ecs/gke": ["Kubernetes"],
        "k8s/gke/helm/docker": ["Kubernetes", "Helm", "Docker"],
        "mysql": ["MySQL"],
        "mysql mongodb": ["MySQL", "MongoDB"],
        "postgres/mssql": ["SQL"],
        "kafka/cassandra/spark bonus": ["Kafka", "Spark"],
        "redis/kafka/spark/k8s bonus": ["Redis", "Kafka", "Spark", "Kubernetes"],
        "spark/kafka/airflow": ["Spark", "Kafka", "Airflow"],
        "spark/redshift/snowflake": ["Spark", "Redshift", "Snowflake"],
        "snowflake/bigquery/redshift": ["Snowflake", "BigQuery", "Redshift"],
        "prometheus/grafana observability": ["Prometheus", "Grafana", "Observability"],
        "grafana/splunk observability": ["Grafana", "Splunk", "Observability"],
        "prometheus/grafana/datadog": ["Prometheus", "Grafana", "Datadog"],
        "prometheus/grafana/loki observability": ["Prometheus", "Grafana", "Observability"],
        "splunk/grafana/new relic": ["Splunk", "Grafana", "New Relic"],
        "pytorch": ["PyTorch"],
        "ml-heavy pytorch/tensorflow": ["PyTorch", "TensorFlow"],
        "wpf ui": ["WPF", "Desktop UI Development"],
        "wpf ui dev": ["WPF", "Desktop UI Development"],
        "wpf or gui frameworks": ["WPF", "Desktop UI Development"],
        "cloud": ["Cloud Infrastructure"],
        "console": ["Console Development"],
        "di": ["Dependency Injection"],
        "di/ioc": ["Dependency Injection", "Inversion of Control"],
        "di/tdd/ci/cd": ["Dependency Injection", "TDD", "CI/CD and Automation"],
        "mvvm": ["MVVM"],
        "mvvm/ioc/di": ["MVVM", "Inversion of Control", "Dependency Injection"],
        "tdd": ["TDD"],
        "tdd/ci": ["TDD", "CI/CD and Automation"],
        "sdlc": ["SDLC"],
        "sdlc qa": ["SDLC"],
        "secure sdlc": ["SDLC"],
        "entity framework": ["Entity Frameworks"],
        "devops": ["DevOps"],
        "llm apis": ["LLM APIs"],
        "llm/genai integration": ["LLM APIs"],
        "genai/llms langchain": ["LLM APIs", "LangChain"],
        "windows internals": ["Windows Internals"],
        "embedded systems": ["Embedded Systems"],
        "embedded linux": ["Embedded Linux"],
        "event-driven systems": ["Event-Driven Systems"],
        "content pipelines": ["Content Pipelines"],
        "game content pipelines": ["Content Pipelines"],
        "data pipelines/streaming": ["Data Pipelines"],
        "data processing pipelines": ["Data Pipelines"],
        "anti-cheat systems": ["Anti-Cheat Systems"],
        "anti-cheat interest": ["Anti-Cheat Systems"],
        "anti-cheat/anti-malware": ["Anti-Cheat Systems", "Anti-Malware"],
        "anti-cheat/fraud": ["Anti-Cheat Systems"],
        "animation pipelines": ["Animation Pipelines"],
        "audio pipelines": ["Audio Pipelines"],
        "physics": ["Physics Systems"],
        "animation": ["Animation Systems"],
        "animation desirable": ["Animation Systems"],
        "physics/animation desirable": ["Physics Systems", "Animation Systems"],
        "physics/animation": ["Physics Systems", "Animation Systems"],
        "gameplay architecture": ["Gameplay Architecture"],
        "multiplayer gameplay": ["Multiplayer Gameplay"],
        "multiplayer server": ["Multiplayer Server Development"],
        "frontend frameworks": ["Frontend Frameworks"],
        "javascript frameworks": ["Frontend Frameworks", "JavaScript"],
        "adtech/monetization systems": ["AdTech Systems"],
        "distributed backend": ["Distributed Backend"],
        "reliability engineering": ["Reliability Engineering"],
        "cloud on-prem": ["Hybrid Infrastructure"],
        "cloud aws/azure/vmware": ["Cloud Infrastructure", "AWS", "Azure", "VMware"],
        "cloud devops": ["Cloud Infrastructure", "DevOps"],
        "toolchain dev": ["Toolchain Development"],
        "frontend tooling": ["Developer Tooling"],
        "developer tooling": ["Developer Tooling"],
        "automation tools": ["Developer Tooling"],
        "code quality": ["Code Quality"],
        "automated testing": ["Test Automation"],
        "automated testing infra": ["Test Automation"],
        "automated testing systems": ["Test Automation"],
        "qa workflows": ["Test Automation"],
        "sdlc qa": ["SDLC", "Test Automation"],
        "etl/data modeling": ["Data Pipelines"],
        "gdpr": ["Data Privacy Compliance"],
        "ocsmp/space/dod preferred": ["Space Technology"],
        "perf optimization": ["Performance Optimization"],
        "memory/perf": ["Performance Optimization"],
        "memory/perf optimization": ["Performance Optimization"],
        "engine optimization": ["Performance Optimization"],
        "cross-platform": ["Cross-Platform Delivery"],
    }
    if normalized in composite_map:
        return composite_map[normalized]
    if normalized in SLASH_PRESERVE_TERMS:
        return [cleaned]
    if re.search(r"[+/]", cleaned):
        return split_generic_composite(cleaned)
    return [cleaned]


def extract_existing_keywords(connection: sqlite3.Connection, row: sqlite3.Row, job_policy: dict[str, object] | None = None) -> list[str]:
    policy = job_policy or load_job_policy()
    blocked = dict(policy.get("keyword_blocklist", {}))
    role_id = row["num"]
    requirements = connection.execute(
        """
        SELECT requirement_name, matched_entity_type, matched_normalized, match_method, source
        FROM role_requirements
        WHERE role_id = ?
        ORDER BY
            CASE source
                WHEN 'jd_text' THEN 0
                WHEN 'tracker_notes' THEN 1
                WHEN 'jd_keywords' THEN 2
                ELSE 3
            END,
            CASE match_method
                WHEN 'exact' THEN 0
                WHEN 'alias' THEN 1
                WHEN 'similar' THEN 2
                WHEN 'new_candidate' THEN 3
                ELSE 4
            END,
            requirement_name
        """,
        (role_id,),
    ).fetchall()
    found: list[str] = []
    seen: set[str] = set()

    for req in requirements:
        raw_terms = split_keyword_terms(req["requirement_name"] or "")
        for term in raw_terms:
            normalized = normalize_text(term)
            if not normalized or normalized in blocked or normalized in seen:
                continue
            if len(normalized) < 3 and normalized not in {"c#", "c++"}:
                continue
            found.append(term)
            seen.add(normalized)
            if len(found) >= 12:
                return found

    basis = " ".join(
        filter(
            None,
            [
                row["role"] or "",
                row["location_text"] if "location_text" in row.keys() else "",
                row["compensation_text"] if "compensation_text" in row.keys() else "",
                row["report"] if "report" in row.keys() else "",
                row["notes"] or "",
            ],
        )
    )
    for keyword in extract_keywords(basis, policy):
        normalized = normalize_text(keyword)
        if not normalized or normalized in seen:
            continue
        found.append(keyword)
        seen.add(normalized)
        if len(found) >= 12:
            break
    return found


def extract_company(text: str) -> str | None:
    lines = cleaned_lines(text)
    patterns = [
        re.compile(r"^company:\s*(.+)$", re.IGNORECASE),
        re.compile(r"^about\s+(.+)$", re.IGNORECASE),
        re.compile(r"^at\s+([A-Z][A-Za-z0-9&.,'() -]{1,60})$"),
    ]
    for line in lines[:25]:
        for pattern in patterns:
            match = pattern.match(line)
            if match:
                return clean_label(match.group(1))
    for index, line in enumerate(lines[:12]):
        if re.search(r"\b(engineer|developer|programmer|architect|analyst|designer|manager)\b", line, re.IGNORECASE) and index > 0:
            previous = lines[index - 1]
            if not re.search(r"\b(remote|hybrid|full[- ]?time|contract)\b", previous, re.IGNORECASE):
                return clean_label(previous)
    return None


def extract_role(text: str) -> str | None:
    for line in cleaned_lines(text)[:20]:
        if re.search(r"\b(engineer|developer|programmer|architect|analyst|designer|manager)\b", line, re.IGNORECASE):
            return clean_label(line)
    return None


def extract_location(text: str) -> str | None:
    for line in cleaned_lines(text)[:30]:
        if re.search(r"\b(remote|hybrid|on-site|onsite)\b", line, re.IGNORECASE):
            return clean_label(line)
        if re.match(r"^location:\s*(.+)$", line, re.IGNORECASE):
            return clean_label(line.split(":", 1)[1])
    return None


def extract_work_model(text: str) -> str | None:
    normalized = normalize_text(text)
    if "remote" in normalized:
        return "remote"
    if "hybrid" in normalized:
        return "hybrid"
    if "on site" in normalized or "onsite" in normalized:
        return "onsite"
    return None


def extract_compensation(text: str) -> dict[str, object]:
    source = text or ""
    range_match = re.search(r"\$([\d.,]+)\s*[Kk]?\s*[–-]\s*\$?([\d.,]+)\s*[Kk]?\s*(USD|CAD)?", source)
    monthly_match = re.search(r"\$([\d.,]+)\s*/\s*mo", source, re.IGNORECASE)
    single_match = re.search(r"\$([\d.,]+)\s*[Kk]?\s*(USD|CAD)?", source)

    def parse_money(raw: str | None) -> float | None:
        if not raw:
            return None
        raw = raw.strip()
        has_k = raw.lower().endswith("k")
        numeric = float(re.sub(r"[^0-9.]", "", raw))
        return numeric * 1000 if has_k else numeric

    if re.search(r"\bno salary\b", source, re.IGNORECASE):
        return {"text": "no salary", "salary_min": None, "salary_max": None, "salary_currency": None, "salary_period": None}
    if range_match:
        return {
            "text": range_match.group(0),
            "salary_min": parse_money(range_match.group(1)),
            "salary_max": parse_money(range_match.group(2)),
            "salary_currency": range_match.group(3) or "USD",
            "salary_period": "yearly",
        }
    if monthly_match:
        value = parse_money(monthly_match.group(1))
        return {"text": monthly_match.group(0), "salary_min": value, "salary_max": value, "salary_currency": "USD", "salary_period": "monthly"}
    if single_match:
        value = parse_money(single_match.group(1))
        return {"text": single_match.group(0), "salary_min": value, "salary_max": value, "salary_currency": single_match.group(2) or "USD", "salary_period": "yearly"}
    return {"text": None, "salary_min": None, "salary_max": None, "salary_currency": None, "salary_period": None}


def extract_jd_id(text: str) -> str | None:
    """Extract a vendor/ATS job ID from raw JD text if present."""
    patterns = [
        # Must be followed by an actual ID token (contains digits, or uppercase+digits mix)
        r"(?:job|requisition|posting|position|opening|reference)\s*(?:id|no|num(?:ber)?|code)?[\s#:.-]+([A-Z0-9][A-Z0-9_-]{3,})",
        r"\bReq(?:uisition)?\s*(?:ID|#|No\.?)[\s:]+([A-Z0-9][A-Z0-9_-]{3,})",
        r"\bID[:\s#]+([A-Z0-9][A-Z0-9_-]{4,})",
        r"\bR(\d{5,})\b",  # Workday-style R123456
    ]
    for pat in patterns:
        match = re.search(pat, text[:3000], re.IGNORECASE)
        if match:
            candidate = match.group(1).strip()
            if re.fullmatch(r"\d{4}", candidate):  # 4-digit year
                continue
            if len(candidate) > 30:
                continue
            # Must contain at least one digit to be a real ID
            if not re.search(r"\d", candidate):
                continue
            return candidate
    return None


def _ensure_jd_id_column(connection: sqlite3.Connection) -> None:
    """Add jd_id column to roles if it doesn't exist (idempotent migration)."""
    cols = {row[1] for row in connection.execute("PRAGMA table_info(roles)").fetchall()}
    if "jd_id" not in cols:
        connection.execute("ALTER TABLE roles ADD COLUMN jd_id TEXT")


def _ensure_archetype_column(connection: sqlite3.Connection) -> None:
    """Add archetype column to roles if it doesn't exist (idempotent migration)."""
    cols = {row[1] for row in connection.execute("PRAGMA table_info(roles)").fetchall()}
    if "archetype" not in cols:
        connection.execute("ALTER TABLE roles ADD COLUMN archetype TEXT")


def _ensure_jd_text_column(connection: sqlite3.Connection) -> None:
    """Add jd_text column to roles if it doesn't exist (idempotent migration)."""
    cols = {row[1] for row in connection.execute("PRAGMA table_info(roles)").fetchall()}
    if "jd_text" not in cols:
        connection.execute("ALTER TABLE roles ADD COLUMN jd_text TEXT")


def _ensure_score_pinned_column(connection: sqlite3.Connection) -> None:
    """Add score_pinned column to roles if it doesn't exist (idempotent migration)."""
    cols = {row[1] for row in connection.execute("PRAGMA table_info(roles)").fetchall()}
    if "score_pinned" not in cols:
        connection.execute("ALTER TABLE roles ADD COLUMN score_pinned INTEGER DEFAULT 0")


def _ensure_found_via_columns(connection: sqlite3.Connection) -> None:
    """Add found_via/apply_method columns, rename via->url if needed, and ensure found_via_sources table exists."""
    cols = {row[1] for row in connection.execute("PRAGMA table_info(roles)").fetchall()}
    if "found_via" not in cols:
        connection.execute("ALTER TABLE roles ADD COLUMN found_via TEXT")
    if "apply_method" not in cols:
        connection.execute("ALTER TABLE roles ADD COLUMN apply_method TEXT")
    # Rename via -> url if needed
    if "via" in cols and "url" not in cols:
        connection.execute("ALTER TABLE roles RENAME COLUMN via TO url")
    elif "via" not in cols and "url" not in cols:
        connection.execute("ALTER TABLE roles ADD COLUMN url TEXT")
    # Ensure found_via_sources table exists
    connection.execute("""
        CREATE TABLE IF NOT EXISTS found_via_sources (
            slug TEXT PRIMARY KEY,
            label TEXT NOT NULL,
            sort_priority INTEGER DEFAULT 0
        )
    """)
    # Seed defaults if empty
    count = connection.execute("SELECT COUNT(*) FROM found_via_sources").fetchone()[0]
    if count == 0:
        defaults = [
            ("linkedin", "LinkedIn", 10), ("indeed", "Indeed", 20),
            ("greenhouse", "Greenhouse", 30), ("glassdoor", "Glassdoor", 40),
            ("hitmarker", "Hitmarker", 50), ("workwithindies", "Work With Indies", 60),
            ("company", "Company site", 70), ("referral", "Referral", 80),
            ("other", "Other", 99),
        ]
        connection.executemany(
            "INSERT OR IGNORE INTO found_via_sources (slug, label, sort_priority) VALUES (?, ?, ?)",
            defaults,
        )


def load_found_via_sources(db_path: Path) -> list[dict]:
    """Return list of {slug, label} from found_via_sources table, ordered by sort_priority."""
    if not db_path.is_file():
        return []
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        rows = conn.execute("SELECT slug, label FROM found_via_sources ORDER BY sort_priority, slug").fetchall()
        return [dict(r) for r in rows]


def prompt_found_via(db_path: Path, prefill: str | None = None) -> str:
    """Prompt user to select or enter a found_via source. Confirm-adds unknown values."""
    if prefill:
        return prefill
    if not sys.stdin.isatty():
        return ""
    sources = load_found_via_sources(db_path)
    print("\n  Where did you find this role?")
    for i, s in enumerate(sources, 1):
        print(f"    {i:>2}. {s['label']} ({s['slug']})")
    print(f"     +. Add new source")
    print()
    while True:
        raw = input("  > ").strip().lower()
        if not raw:
            continue
        # add new
        if raw in ("+", "new", "add"):
            slug = input("  Slug (e.g. 'hitmarker'): ").strip().lower()
            if not slug:
                print("  Cancelled.")
                continue
            label = input(f"  Label (or Enter to use '{slug.title()}'): ").strip() or slug.title()
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    "INSERT OR IGNORE INTO found_via_sources (slug, label, sort_priority) VALUES (?, ?, ?)",
                    (slug, label, 50),
                )
                conn.commit()
            print(f"  Added '{slug}'.")
            sources = load_found_via_sources(db_path)
            return slug
        # numeric shortcut
        if raw.isdigit():
            idx = int(raw) - 1
            if 0 <= idx < len(sources):
                return sources[idx]["slug"]
            print(f"  Invalid number. Enter 1\u2013{len(sources)}, + to add new, or type a slug.")
            continue
        # known slug
        known_slugs = {s["slug"] for s in sources}
        if raw in known_slugs:
            return raw
        # unknown slug typed directly — confirm add
        ans = input(f"  '{raw}' not in list. Add it? [y/N] ").strip().lower()
        if ans == "y":
            label = input(f"  Label (or Enter to use '{raw.title()}'): ").strip() or raw.title()
            with sqlite3.connect(db_path) as conn:
                conn.execute(
                    "INSERT OR IGNORE INTO found_via_sources (slug, label, sort_priority) VALUES (?, ?, ?)",
                    (raw, label, 50),
                )
                conn.commit()
            print(f"  Added '{raw}'.")
            return raw
        print("  Try again.")


APPLY_METHODS = [("company", "Company site / ATS"), ("linkedin", "LinkedIn Easy Apply"), ("email", "Email"), ("other", "Other")]


def prompt_apply_method(prefill: str | None = None) -> str:
    """Prompt user to select apply_method."""
    if prefill:
        return prefill
    if not sys.stdin.isatty():
        return ""
    print("\n  How did you apply?")
    for i, (slug, label) in enumerate(APPLY_METHODS, 1):
        print(f"    {i}. {label} ({slug})")
    print()
    while True:
        raw = input("  > ").strip().lower()
        if raw.isdigit():
            idx = int(raw) - 1
            if 0 <= idx < len(APPLY_METHODS):
                return APPLY_METHODS[idx][0]
        if raw in {s for s, _ in APPLY_METHODS}:
            return raw
        print(f"  Enter 1\u2013{len(APPLY_METHODS)} or one of: company, linkedin, email, other")


def log_to_db(repo_root: Path, record: dict[str, object], ai_provider: str, minimal: bool = False) -> tuple[int, list[str]]:
    db_path = repo_root / "data" / "job-log.db"
    company = str(record["company"]).strip()
    role = str(record["role"]).strip()
    location = str(record.get("location") or "").strip()
    work_model = str(record.get("work_model") or "").strip() or extract_work_model(location or str(record.get("jd_text") or ""))
    compensation = dict(record.get("compensation") or {})
    source = dict(record["source"])
    jd_text = str(record.get("jd_text") or "")
    archetype = str(record["archetype"]).strip()
    keywords = [str(keyword).strip() for keyword in record.get("keywords", []) if str(keyword).strip()]
    notes = f"ai={ai_provider}; source={source['type']}"
    today = today_iso()

    jd_id = record.get("jd_id") or (extract_jd_id(jd_text) if jd_text else None)
    initial_status = record.get("initial_status") or "Evaluated"

    found_via = str(record.get("found_via") or "").strip() or None
    apply_method = str(record.get("apply_method") or "").strip() or None

    with sqlite3.connect(db_path) as connection:
        connection.row_factory = sqlite3.Row
        _ensure_jd_id_column(connection)
        _ensure_archetype_column(connection)
        _ensure_jd_text_column(connection)
        _ensure_score_pinned_column(connection)
        _ensure_found_via_columns(connection)
        next_num = connection.execute("SELECT COALESCE(MAX(num), 0) + 1 FROM roles").fetchone()[0]
        derived_score = derive_role_score_from_text(role, jd_text, compensation, work_model, keywords, load_job_policy())
        connection.execute(
            """
            INSERT INTO roles (
                num, date, first_seen_date, last_updated_date, company, role,
                location_text, work_model, compensation_text, salary_min, salary_max, salary_currency, salary_period,
                score, status, notes, source, url, jd_id, archetype, jd_text, found_via, apply_method
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(company, role) DO UPDATE SET
                last_updated_date = excluded.last_updated_date,
                location_text = COALESCE(excluded.location_text, roles.location_text),
                work_model = COALESCE(excluded.work_model, roles.work_model),
                compensation_text = COALESCE(excluded.compensation_text, roles.compensation_text),
                salary_min = COALESCE(excluded.salary_min, roles.salary_min),
                salary_max = COALESCE(excluded.salary_max, roles.salary_max),
                salary_currency = COALESCE(excluded.salary_currency, roles.salary_currency),
                salary_period = COALESCE(excluded.salary_period, roles.salary_period),
                score = COALESCE(roles.score, excluded.score),
                notes = excluded.notes,
                source = COALESCE(excluded.source, roles.source),
                url = COALESCE(excluded.url, roles.url),
                jd_id = COALESCE(excluded.jd_id, roles.jd_id),
                archetype = excluded.archetype,
                jd_text = COALESCE(excluded.jd_text, roles.jd_text),
                found_via = COALESCE(excluded.found_via, roles.found_via),
                apply_method = COALESCE(excluded.apply_method, roles.apply_method)
            """,
            (
                next_num,
                today,
                today,
                today,
                company,
                role,
                location or None,
                work_model or None,
                compensation.get("text"),
                compensation.get("salary_min"),
                compensation.get("salary_max"),
                compensation.get("salary_currency"),
                compensation.get("salary_period"),
                derived_score,
                initial_status,
                notes,
                source.get("type"),
                source.get("value"),
                jd_id,
                archetype,
                jd_text or None,
                found_via,
                apply_method,
            ),
        )
        role_row = connection.execute("SELECT num FROM roles WHERE company = ? AND role = ?", (company, role)).fetchone()
        if role_row is None:
            raise SystemExit("Role insert failed.")
        # If an explicit non-default status was requested, force-update even on conflict
        if initial_status != "Evaluated":
            connection.execute(
                "UPDATE roles SET status = ?, last_updated_date = ? WHERE num = ?",
                (initial_status, today, role_row["num"]),
            )
        new_candidates = [] if minimal else write_role_requirements(
            connection,
            role_row["num"],
            role,
            jd_text,
            keywords,
            record.get("requirements"),
        )
        upsert_skill_signals(connection, role_row["num"], archetype)
        connection.commit()
        return int(role_row["num"]), new_candidates


def load_matcher_rows(connection: sqlite3.Connection) -> list[dict[str, object]]:
    rows: list[dict[str, object]] = []
    for kind, canonical_table, name_col, normalized_col, alias_table, alias_fk in (
        ("skill", "skills_mine", "skill_name", "skill_normalized", "skill_aliases", "skill_normalized"),
        ("capability", "capabilities_mine", "capability_name", "capability_normalized", "capability_aliases", "capability_normalized"),
        ("qualification", "qualifications_mine", "qualification_name", "qualification_normalized", "qualification_aliases", "qualification_normalized"),
    ):
        for row in connection.execute(f"SELECT {name_col} AS name, {normalized_col} AS normalized FROM {canonical_table}").fetchall():
            rows.append({"kind": kind, "term": row["name"], "normalized": row["normalized"], "matched_normalized": row["normalized"], "match_method": "exact"})
        for row in connection.execute(f"SELECT alias_name AS name, alias_normalized AS normalized, {alias_fk} AS matched_normalized FROM {alias_table}").fetchall():
            rows.append({"kind": kind, "term": row["name"], "normalized": row["normalized"], "matched_normalized": row["matched_normalized"], "match_method": "alias"})
    rows.sort(key=lambda item: len(str(item["normalized"])), reverse=True)
    return rows


def should_skip_inventory_candidate(name: str, normalized: str) -> bool:
    if not normalized:
        return True
    if re.fullmatch(r"\d+\+?\s*yrs?(?:\s+req)?", normalized):
        return True
    if re.fullmatch(r"\d+(?:[–-]\d+)?\s*yrs?(?:\s+req)?", normalized):
        return True
    if re.search(r"\b(?:travel|salary|application|decision|match|fit|source|resume|contact|interview|portfolio|recruiter|linkedin)\b", normalized):
        return True
    # Don't skip clearance/location/availability — these are handled by qualifications_mine

    if re.search(r"\bgaps?\b", normalized):
        return True
    if re.search(r"\b(?:relevant|connection noted|old company|reached out|desired|optional)\b", normalized):
        return True
    if re.search(r"applying despite overqualification", normalized):
        return True
    if re.search(r"game dev desirable", normalized):
        return True
    if re.match(r"^(?:same as|used to play|played |no named |no portfolio |no foreign |lighter exp)", normalized):
        return True
    return False


def infer_skill_category(name: str, normalized: str) -> tuple[str, str, str | None]:
    if normalized in {"aws", "gcp", "azure"} or any(token in normalized for token in ["k8s", "kubernetes", "helm", "terraform"]):
        return "cloud", "Cloud", "Cloud / Backend"
    if normalized in {"go", "typescript", "javascript", "kotlin"}:
        return "language", "Languages", None
    if any(token in normalized for token in ["junit", "mockito", "maven", "gradle", "jenkins", "gitlab", "github actions", "cloud build"]):
        return "tool", "Tools", "Developer Tools"
    if any(token in normalized for token in ["websocket", "grpc", "rest", "api", "sql", "nosql", "database"]):
        return "backend", "Backend", "Cloud / Backend"
    if any(token in normalized for token in ["docker", "perforce", "git", "linux"]):
        return "tool", "Tools", "Developer Tools"
    return "tool", "Tools", None


def infer_capability_category(normalized: str) -> str:
    if any(token in normalized for token in ["monitor", "observability", "analytics", "telemetry"]):
        return "observability"
    if any(token in normalized for token in ["architecture", "scalability", "distributed", "microservices", "systems", "multithreading"]):
        return "systems"
    if any(token in normalized for token in ["incident", "response", "operations"]):
        return "operations"
    if any(token in normalized for token in ["graphics"]):
        return "graphics"
    if any(token in normalized for token in ["console", "platform"]):
        return "platforms"
    if any(token in normalized for token in ["gitops", "slo", "sli"]):
        return "devops" if "gitops" in normalized else "observability"
    if any(token in normalized for token in ["testing", "troubleshooting", "debug", "optimization"]):
        return "engineering"
    if any(token in normalized for token in ["pipelines", "data"]):
        return "data"
    return "systems"


def candidate_should_be_capability(normalized: str) -> bool:
    if "/" in normalized and any(token in normalized for token in ["aws", "gcp", "azure", "docker", "k8s", "sql", "rest", "grpc", "websocket", "c#", "c++", "java", "python", "go", "javascript", "typescript"]):
        return False
    if re.search(r"\b(?:architecture|scalability|distributed|microservices|multithreading|observability|incident response|data pipelines|unit testing|debugging|troubleshooting|optimization|integrity|graphics apis|cloud infrastructure|gitops|slo/sli|console development|database systems|oop|build systems|development workflows|large codebase maintenance|latency mitigation|low-level systems|technical mentoring|networking|rendering systems|gpu architecture|agile)\b", normalized):
        return True
    return " " in normalized and normalized.islower()


def upsert_inventory_candidate(connection: sqlite3.Connection, requirement_name: str) -> None:
    cleaned = clean_label(requirement_name)
    normalized = normalize_text(cleaned)
    if should_skip_inventory_candidate(cleaned, normalized):
        return
    if candidate_should_be_capability(normalized):
        connection.execute(
            """
            INSERT INTO capabilities_mine (capability_name, capability_normalized, category, level, resume_priority, evidence, notes)
            VALUES (?, ?, ?, 'none', 1, 'Auto-added from role requirement candidate; no evidence recorded yet', 'Auto-added from role upsert candidate review')
            ON CONFLICT(capability_normalized) DO NOTHING
            """,
            (cleaned, normalized, infer_capability_category(normalized)),
        )
        return

    category, display_category, secondary_categories = infer_skill_category(cleaned, normalized)
    connection.execute(
        """
        INSERT INTO skills_mine (
            skill_name, skill_normalized, category, display_category, secondary_categories,
            resume_priority, include_default, require_direct_match, profile_bias,
            resume_visibility, resume_display, resume_emphasis, resume_group_rank, level, evidence
        )
        VALUES (?, ?, ?, ?, ?, 1, 0, 1, NULL, 'context', NULL, 'plain', 0, 'none', 'Auto-added from role requirement candidate; no evidence recorded yet')
        ON CONFLICT(skill_normalized) DO NOTHING
        """,
        (cleaned, normalized, category, display_category, secondary_categories),
    )


def collect_requirement_candidates(requirements: list[tuple[object, ...]]) -> list[str]:
    """Return cleaned candidate names from new_candidate requirements, skipping junk terms."""
    candidates: list[str] = []
    seen: set[str] = set()
    for requirement in requirements:
        if len(requirement) < 10:
            continue
        if str(requirement[8] or "") != "new_candidate":
            continue
        for term in expand_requirement_term(str(requirement[2] or "")):
            cleaned = clean_label(term)
            normalized = normalize_text(cleaned)
            if should_skip_inventory_candidate(cleaned, normalized) or normalized in seen:
                continue
            seen.add(normalized)
            candidates.append(cleaned)
    return candidates


def prompt_new_inventory_items(candidates: list[str], db_path: Path) -> None:
    """Interactively prompt user to categorize new JD terms not yet in their inventory."""
    if not candidates or not sys.stdin.isatty():
        return

    LEVEL_MAP = {"0": "none", "1": "exposure", "2": "basic", "3": "intermediate", "4": "advanced", "5": "expert"}
    MET_MAP = {"met": "yes", "yes": "yes", "unmet": "no", "no": "no", "partial": "partial"}

    def _similar(norm: str, inv: list[dict], query: str = "", limit: int = 5) -> list[dict]:
        search = normalize_text(query) if query else norm
        s_tokens = set(search.split())
        scored = []
        for entry in inv:
            en = entry["normalized"]
            en_tokens = set(en.split())
            if s_tokens <= en_tokens or en_tokens <= s_tokens:
                score = 0.85
            else:
                score = difflib.SequenceMatcher(None, search, en).ratio()
                if query and any(t in en for t in s_tokens):
                    score = max(score, 0.6)
            if score >= 0.55:
                scored.append((score, entry))
        scored.sort(key=lambda x: -x[0])
        return [e for _, e in scored[:limit]]

    print(f"\n  {len(candidates)} new term(s) found — not in your inventory.")
    print("  [s]kill / [c]apability / [q]ualification / s[k]ip  —  s 4 · c 3 · q met · ?search\n")

    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row

        inventory: list[dict] = []
        for row in conn.execute("SELECT skill_name, skill_normalized, level FROM skills_mine"):
            inventory.append({"kind": "skill", "name": row[0], "normalized": row[1], "detail": row[2] or "none"})
        for row in conn.execute("SELECT capability_name, capability_normalized, level FROM capabilities_mine"):
            inventory.append({"kind": "cap", "name": row[0], "normalized": row[1], "detail": row[2] or "none"})
        for row in conn.execute("SELECT qualification_name, qualification_normalized, met FROM qualifications_mine"):
            inventory.append({"kind": "qual", "name": row[0], "normalized": row[1], "detail": row[2] or "?"})

        for term in candidates:
            normalized = normalize_text(clean_label(term))
            if (
                conn.execute("SELECT 1 FROM skills_mine WHERE skill_normalized = ?", (normalized,)).fetchone()
                or conn.execute("SELECT 1 FROM skill_aliases WHERE alias_normalized = ?", (normalized,)).fetchone()
                or conn.execute("SELECT 1 FROM capabilities_mine WHERE capability_normalized = ?", (normalized,)).fetchone()
                or conn.execute("SELECT 1 FROM capability_aliases WHERE alias_normalized = ?", (normalized,)).fetchone()
                or conn.execute("SELECT 1 FROM qualifications_mine WHERE qualification_normalized = ?", (normalized,)).fetchone()
            ):
                continue

            base_matches = _similar(normalized, inventory)
            extended_matches = _similar(normalized, inventory, limit=15)
            current_matches = base_matches

            while True:
                print(f"  '{term}'", end="")
                if current_matches:
                    print(" — similar:")
                    for i, m in enumerate(current_matches, 1):
                        kind_label = {"skill": "skill", "cap": "capability", "qual": "qualification"}[m["kind"]]
                        print(f"    {i}. {m['name']} ({kind_label}, {m['detail']})")
                    if current_matches is base_matches and len(extended_matches) > len(base_matches):
                        print(f"    ({len(extended_matches) - len(base_matches)} more — use ? to expand)")
                else:
                    print()
                print("  [s]kill (0-5) [name] · [c]apability (0-5) [name] · [q]ualification [met] · #N link to match · ?search · s[k]ip")

                raw_orig = input("  → ").strip()
                raw = raw_orig.lower()
                print()

                if not raw or raw in ("k", "sk", "skip"):
                    print("    → skipped\n")
                    break

                if raw.startswith("?"):
                    query = raw[1:].strip()
                    if query:
                        current_matches = _similar(normalized, inventory, query)
                    else:
                        current_matches = extended_matches
                    if not current_matches:
                        print("    no matches\n")
                    continue

                parts = raw.split(None, 2)
                parts_orig = raw_orig.split(None, 2)
                cmd = parts[0]
                arg1 = parts[1] if len(parts) > 1 else ""
                arg2 = parts[2] if len(parts) > 2 else ""
                arg1_orig = parts_orig[1] if len(parts_orig) > 1 else ""
                arg2_orig = parts_orig[2] if len(parts_orig) > 2 else ""

                if cmd.isdigit():
                    idx = int(cmd) - 1
                    if 0 <= idx < len(current_matches):
                        match = current_matches[idx]
                        # arg1: level (1-5 or name) → also update level; otherwise custom display name
                        level_str = LEVEL_MAP.get(arg1) or (arg1 if arg1 in LEVEL_MAP.values() else None)
                        alias_display = arg2_orig if (level_str and arg2_orig) else (arg1_orig if (arg1_orig and not level_str) else term)
                        if match["kind"] == "qual":
                            print(f"    → covered by '{match['name']}' (qualification)\n")
                            break
                        warn = len(normalized.split()) == 1 and len(normalized) <= 5
                        scope_note = "  ! short/generic term — will match ALL future JDs\n" if warn else "  (will match all future JDs)\n"
                        level_note = f" + level={level_str}" if level_str else ""
                        canon_name = match["name"]
                        print(f"    link '{alias_display}' → '{canon_name}' ({match['kind']}){level_note}")
                        print(f"  {scope_note}", end="")
                        confirm = input("  save? [Y/n/r rename/l level]: ").strip().lower()
                        print()
                        if confirm in ("n", "no"):
                            print("    → cancelled\n")
                            break
                        if confirm.startswith("l"):
                            edited_level = input(f"  level (0-5 or name) [{level_str or 'unchanged'}]: ").strip().lower()
                            if edited_level in LEVEL_MAP:
                                level_str = LEVEL_MAP[edited_level]
                            elif edited_level in LEVEL_MAP.values():
                                level_str = edited_level
                            print()
                        if confirm.startswith("r"):
                            new_name = input(f"  rename '{canon_name}' to: ").strip()
                            print()
                            if new_name:
                                canon_table = "skills_mine" if match["kind"] == "skill" else "capabilities_mine"
                                name_col = "skill_name" if match["kind"] == "skill" else "capability_name"
                                norm_col = "skill_normalized" if match["kind"] == "skill" else "capability_normalized"
                                conn.execute(
                                    f"UPDATE {canon_table} SET {name_col} = ? WHERE {norm_col} = ?",
                                    (new_name, match["normalized"]),
                                )
                                canon_name = new_name
                                print(f"    → renamed to '{canon_name}'\n")
                        table = "skill_aliases" if match["kind"] == "skill" else "capability_aliases"
                        fk_col = "skill_normalized" if match["kind"] == "skill" else "capability_normalized"
                        conn.execute(
                            f"INSERT INTO {table} ({fk_col}, alias_name, alias_normalized, notes) "
                            f"VALUES (?, ?, ?, 'Added via intake prompt') ON CONFLICT(alias_normalized) DO NOTHING",
                            (match["normalized"], alias_display, normalized),
                        )
                        if level_str:
                            canon_table = "skills_mine" if match["kind"] == "skill" else "capabilities_mine"
                            norm_col = "skill_normalized" if match["kind"] == "skill" else "capability_normalized"
                            conn.execute(
                                f"UPDATE {canon_table} SET level = ? WHERE {norm_col} = ?",
                                (level_str, match["normalized"]),
                            )
                        conn.commit()
                        level_saved = f", level={level_str}" if level_str else ""
                        print(f"    → linked{level_saved}\n")
                        break
                    else:
                        print(f"    ? pick 1–{len(current_matches)}\n")
                    continue

                if cmd in ("s", "skill"):
                    if arg1 in LEVEL_MAP:
                        level, custom_name = LEVEL_MAP[arg1], arg2_orig
                    else:
                        level, custom_name = "none", (" ".join(parts_orig[1:]) if arg1 else "")
                    skill_name = custom_name if custom_name else term
                    skill_norm = normalize_text(skill_name)
                    cat, disp_cat, sec_cats = infer_skill_category(skill_name, skill_norm)
                    confirm = input(f"  save skill '{skill_name}' [{level}]? [Y/n/e to edit]: ").strip().lower()
                    print()
                    if confirm in ("n", "no"):
                        print("    → cancelled\n")
                        continue
                    if confirm.startswith("e"):
                        edited = input(f"  name [{skill_name}]: ").strip()
                        if edited:
                            skill_name = edited
                            skill_norm = normalize_text(skill_name)
                            cat, disp_cat, sec_cats = infer_skill_category(skill_name, skill_norm)
                        edited_level = input(f"  level [{level}]: ").strip().lower()
                        if edited_level in LEVEL_MAP.values():
                            level = edited_level
                        elif edited_level in LEVEL_MAP:
                            level = LEVEL_MAP[edited_level]
                        print()
                    conn.execute(
                        """
                        INSERT INTO skills_mine (skill_name, skill_normalized, category, display_category, secondary_categories,
                            resume_priority, include_default, require_direct_match, profile_bias,
                            resume_visibility, resume_display, resume_emphasis, resume_group_rank, level, evidence)
                        VALUES (?, ?, ?, ?, ?, 1, 0, 1, NULL, 'context', NULL, 'plain', 0, ?, 'Added via intake prompt')
                        ON CONFLICT(skill_normalized) DO UPDATE SET level = excluded.level
                        """,
                        (skill_name, skill_norm, cat, disp_cat, sec_cats, level),
                    )
                    if custom_name and skill_norm != normalized:
                        warn = len(normalized.split()) == 1 and len(normalized) <= 5
                        scope_note = "! short/generic term — will match ALL future JDs" if warn else "will match all future JDs"
                        confirm = input(f"  also alias '{term}' → '{skill_name}' globally? ({scope_note}) [Y/n]: ").strip().lower()
                        print()
                        if confirm not in ("n", "no"):
                            conn.execute(
                                "INSERT INTO skill_aliases (skill_normalized, alias_name, alias_normalized, notes) "
                                "VALUES (?, ?, ?, 'JD term aliased at intake') ON CONFLICT(alias_normalized) DO NOTHING",
                                (skill_norm, term, normalized),
                            )
                    conn.commit()
                    print(f"    → skill '{skill_name}' [{level}]\n")
                    break
                elif cmd in ("c", "capability"):
                    if arg1 in LEVEL_MAP:
                        level, custom_name = LEVEL_MAP[arg1], arg2_orig
                    else:
                        level, custom_name = "none", (" ".join(parts_orig[1:]) if arg1 else "")
                    cap_name = custom_name if custom_name else term
                    cap_norm = normalize_text(cap_name)
                    cat = infer_capability_category(cap_norm)
                    confirm = input(f"  save capability '{cap_name}' [{level}]? [Y/n/e to edit]: ").strip().lower()
                    print()
                    if confirm in ("n", "no"):
                        print("    → cancelled\n")
                        continue
                    if confirm.startswith("e"):
                        edited = input(f"  name [{cap_name}]: ").strip()
                        if edited:
                            cap_name = edited
                            cap_norm = normalize_text(cap_name)
                            cat = infer_capability_category(cap_norm)
                        edited_level = input(f"  level [{level}]: ").strip().lower()
                        if edited_level in LEVEL_MAP.values():
                            level = edited_level
                        elif edited_level in LEVEL_MAP:
                            level = LEVEL_MAP[edited_level]
                        print()
                    conn.execute(
                        """
                        INSERT INTO capabilities_mine (capability_name, capability_normalized, category, level, resume_priority, evidence, notes)
                        VALUES (?, ?, ?, ?, 1, 'Added via intake prompt', NULL)
                        ON CONFLICT(capability_normalized) DO UPDATE SET level = excluded.level
                        """,
                        (cap_name, cap_norm, cat, level),
                    )
                    if custom_name and cap_norm != normalized:
                        warn = len(normalized.split()) == 1 and len(normalized) <= 5
                        scope_note = "! short/generic term — will match ALL future JDs" if warn else "will match all future JDs"
                        confirm = input(f"  also alias '{term}' → '{cap_name}' globally? ({scope_note}) [Y/n]: ").strip().lower()
                        print()
                        if confirm not in ("n", "no"):
                            conn.execute(
                                "INSERT INTO capability_aliases (capability_normalized, alias_name, alias_normalized, notes) "
                                "VALUES (?, ?, ?, 'JD term aliased at intake') ON CONFLICT(alias_normalized) DO NOTHING",
                                (cap_norm, term, normalized),
                            )
                    conn.commit()
                    print(f"    → capability '{cap_name}' [{level}]\n")
                    break
                elif cmd in ("q", "qualification"):
                    met = MET_MAP.get(arg1, "no")
                    confirm = input(f"  save qualification '{term}' [met={met}]? [Y/n/e to edit]: ").strip().lower()
                    print()
                    if confirm in ("n", "no"):
                        print("    → cancelled\n")
                        continue
                    if confirm.startswith("e"):
                        edited_met = input(f"  met [{met}] (yes/no/partial): ").strip().lower()
                        if edited_met in MET_MAP:
                            met = MET_MAP[edited_met]
                        elif edited_met in ("yes", "no", "partial"):
                            met = edited_met
                        print()
                    conn.execute(
                        """
                        INSERT INTO qualifications_mine (qualification_name, qualification_normalized, category, met, notes)
                        VALUES (?, ?, 'experience', ?, 'Added via intake prompt')
                        ON CONFLICT(qualification_normalized) DO NOTHING
                        """,
                        (term, normalized, met),
                    )
                    conn.commit()
                    print(f"    → qualification [met={met}]\n")
                    break
                else:
                    print("    [s]kill (0-5) [name] · [c]apability (0-5) [name] · [q]ualification [met] · #N link to match · ?search · s[k]ip\n")


def upsert_skill_signals(connection: sqlite3.Connection, role_id: int, archetype: str) -> None:
    arch_name = DB_ARCHETYPE_NAMES.get(archetype, archetype)
    connection.execute("INSERT OR IGNORE INTO archetypes (name, evals) VALUES (?, 0)", (arch_name,))
    connection.execute("UPDATE archetypes SET evals = evals + 1 WHERE name = ?", (arch_name,))
    arc_id = connection.execute("SELECT id FROM archetypes WHERE name = ?", (arch_name,)).fetchone()
    if arc_id is None:
        return
    arc_id = arc_id[0]

    my_skills = {r[0] for r in connection.execute("SELECT skill_normalized FROM skills_mine").fetchall()}
    reqs = connection.execute(
        "SELECT requirement_name, requirement_normalized, matched_normalized FROM role_requirements WHERE role_id = ?",
        (role_id,),
    ).fetchall()

    upsert = """
        INSERT INTO skill_signals (archetype_id, type, skill_name, skill_normalized, count)
        VALUES (?, ?, ?, ?, 1)
        ON CONFLICT(archetype_id, type, skill_normalized) DO UPDATE SET count = count + 1
    """
    for req in reqs:
        matched = req[2]
        skill_normalized = matched if matched else req[1]
        signal_type = "matched" if matched and matched in my_skills else "missing"
        connection.execute(upsert, (arc_id, signal_type, req[0], skill_normalized))


def empty_requirement_groups() -> dict[str, list[str]]:
    return {"skills": [], "capabilities": [], "qualifications": []}


def normalize_requirement_groups(value: object) -> dict[str, list[str]]:
    groups = empty_requirement_groups()
    if not isinstance(value, dict):
        return groups
    for key in groups:
        groups[key] = normalize_ai_string_list(value.get(key) or [], max_items=20, label=key)
    return groups


def add_requirement_entry(
    requirements: list[tuple[object, ...]],
    seen: set[tuple[str, str]],
    *,
    role_id: int,
    raw_text: str,
    requirement_name: str,
    kind: str,
    matcher_by_norm: dict[str, dict],
    source: str,
    notes: str | None = None,
    confidence: float | None = None,
    match_method_override: str | None = None,
) -> None:
    normalized = normalize_text(requirement_name)
    if not normalized:
        return
    key = (kind, normalized)
    if key in seen:
        return
    seen.add(key)
    match = matcher_by_norm.get(normalized)
    matched_kind = str(match["kind"]) if match else None
    matched_normalized = str(match["matched_normalized"]) if match else None
    matched_entity_type = matched_kind if match and matched_kind in {"skill", "capability", "qualification"} else "none"
    match_method = match_method_override or (str(match["match_method"]) if match else "new_candidate")
    requirements.append(
        (
            role_id,
            raw_text,
            requirement_name,
            normalized,
            kind,
            "unknown",
            matched_entity_type,
            matched_normalized,
            match_method,
            confidence,
            source,
            notes,
        )
    )


def write_role_requirements(
    connection: sqlite3.Connection,
    role_id: int,
    role: str,
    jd_text: str,
    keywords: list[str],
    requirement_groups: dict[str, list[str]] | None = None,
) -> list[str]:
    basis = normalize_text(f"{role} {jd_text}")
    matcher_rows = load_matcher_rows(connection)
    # Index matcher_rows by normalized value for O(1) keyword lookup
    matcher_by_norm: dict[str, dict] = {}
    for item in matcher_rows:
        n = str(item["normalized"]).strip()
        if n and n not in matcher_by_norm:
            matcher_by_norm[n] = item
    seen: set[tuple[str, str]] = set()
    requirements: list[tuple[object, ...]] = []
    groups = normalize_requirement_groups(requirement_groups)

    for item in matcher_rows:
        normalized = str(item["normalized"]).strip()
        if not normalized:
            continue
        pattern = re.compile(rf"(^| ){re.escape(normalized)}( |$)")
        if not pattern.search(basis):
            continue
        key = (str(item["kind"]), str(item["matched_normalized"]))
        if key in seen:
            continue
        seen.add(key)
        requirements.append(
            (
                role_id,
                str(item["term"]),
                str(item["term"]),
                normalized,
                str(item["kind"]),
                "unknown",
                str(item["kind"]),
                str(item["matched_normalized"]),
                str(item["match_method"]),
                1.0 if item["match_method"] == "exact" else 0.9,
                "jd_text",
                None,
            )
        )

    for group_key, kind in (("skills", "skill"), ("capabilities", "capability"), ("qualifications", "qualification")):
        for term in groups.get(group_key, []):
            add_requirement_entry(
                requirements,
                seen,
                role_id=role_id,
                raw_text=term,
                requirement_name=term,
                kind=kind,
                matcher_by_norm=matcher_by_norm,
                source="ai_grouped",
                notes="AI-grouped requirement",
                confidence=0.95,
            )

    for keyword in keywords:
        for expanded in expand_requirement_term(keyword):
            normalized = normalize_text(expanded)
            if not normalized:
                continue
            match = matcher_by_norm.get(normalized)
            ai_kind = None
            for group_key, kind_name in (("skills", "skill"), ("capabilities", "capability"), ("qualifications", "qualification")):
                if normalized in {normalize_text(item) for item in groups.get(group_key, [])}:
                    ai_kind = kind_name
                    break
            kind = ai_kind or (str(match["kind"]) if match else "unknown")
            key = (kind, normalized)
            if key in seen:
                continue
            if any(normalized == row[3] for row in requirements):
                continue
            if match:
                requirements.append(
                    (
                        role_id,
                        keyword,
                        expanded,
                        normalized,
                        kind,
                        "unknown",
                        kind,
                        str(match["matched_normalized"]),
                        str(match["match_method"]),
                        1.0 if match["match_method"] == "exact" else 0.9,
                        "jd_keywords",
                        None,
                    )
                )
            else:
                requirements.append(
                    (
                        role_id,
                        keyword,
                        expanded,
                        normalized,
                        "unknown",
                        "unknown",
                        "none",
                        None,
                        "new_candidate",
                        0.4,
                        "jd_keywords",
                        "Unmatched extracted keyword; review as possible alias or new inventory entry.",
                    )
                )
            seen.add(key)

    connection.execute("DELETE FROM role_requirements WHERE role_id = ?", (role_id,))
    connection.executemany(
        """
        INSERT INTO role_requirements (
            role_id, raw_text, requirement_name, requirement_normalized, kind, priority,
            matched_entity_type, matched_normalized, match_method, confidence, source, notes
        ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
        ON CONFLICT(role_id, requirement_normalized, kind, priority) DO UPDATE SET
            raw_text = excluded.raw_text,
            requirement_name = excluded.requirement_name,
            matched_entity_type = excluded.matched_entity_type,
            matched_normalized = excluded.matched_normalized,
            match_method = excluded.match_method,
            confidence = excluded.confidence,
            source = excluded.source,
            notes = excluded.notes
        """,
        requirements,
    )
    return collect_requirement_candidates(requirements)


def derive_role_score_from_text(
    role: str,
    jd_text: str,
    compensation: dict[str, object],
    work_model: str | None,
    keywords: list[str],
    job_policy: dict[str, object] | None = None,
) -> float:
    text = f"{role} {jd_text}"
    lower = text.lower()
    classification = classify_job_description(text, job_policy or load_job_policy(), include_details=True)
    matched_weight = sum(match["weight"] for match in classification["analysis"]["matched_rules"])
    seniority_penalty = 0.5 if re.search(r"\b(lead|principal|staff|manager|director)\b", role, re.IGNORECASE) else 0.0
    fit_bonus = archetype_fit_modifier(classification["archetype"])
    salary_max = float(compensation.get("salary_max") or 0)
    salary_min = float(compensation.get("salary_min") or 0)
    if salary_max > 0 and salary_max <= 80000:
        fit_bonus -= 2.0  # At or below minimum threshold
    elif salary_max > 0 and salary_max < 90000:
        fit_bonus -= 1.0  # Below comfortable range
    elif salary_min >= 100000:
        fit_bonus += 0.5
    if work_model == "remote":
        fit_bonus += 0.4
    elif work_model == "hybrid":
        fit_bonus += 0.1
    elif work_model == "onsite":
        fit_bonus -= 0.3
    if "security clearance" in lower or "secret clearance" in lower or "ts/sci" in lower:
        fit_bonus -= 0.5
    # Long contract at low salary: if JD is a contract role (not permanent) and pay is below threshold
    is_contract = bool(re.search(r'\b(contract|contractor|freelance|temp|temporary)\b', lower))
    if is_contract and salary_max > 0 and salary_max <= 90000:
        fit_bonus -= 0.5  # Additional penalty: long-term commitment at low pay
    raw_score = 2.5 + min(3.5, matched_weight / 5) + min(1.5, len(keywords) / 6) + fit_bonus - seniority_penalty
    return normalize_role_score(raw_score)


def _python_executable() -> str:
    """Return a Python executable that works in subprocesses on Windows.

    On Windows, sys.executable may be an AppX Store shim that cannot be
    launched as a regular child process. Prefer the 'py' launcher (py.exe)
    which is always a real executable, then fall back to sys.executable.
    """
    if sys.platform == "win32":
        py_launcher = shutil.which("py")
        if py_launcher:
            return py_launcher
    return sys.executable


def run_resume_generation(repo_root: Path, record: dict[str, object], output_dir: Path) -> Path:
    script = repo_root / "scripts" / "generate-resume.py"
    if not script.is_file():
        raise SystemExit(f"Resume script not found: {script}")

    archetype = str(record["archetype"]).strip()
    command = [
        _python_executable(),
        str(script),
        f"--archetype={archetype}",
        f"--keywords={','.join(str(keyword).strip() for keyword in record.get('keywords', []) if str(keyword).strip())}",
        f"--job-role={str(record.get('role') or '').strip()}",
        f"--job-company={str(record.get('company') or '').strip()}",
        f"--job-location={str(record.get('location_text') or record.get('location') or '').strip()}",
        f"--job-work-model={str(record.get('work_model') or '').strip()}",
        f"--job-compensation={str(record.get('compensation_text') or '').strip()}",
        f"--out={output_dir}",
    ]
    run_subprocess(command, cwd=repo_root)

    target_pdf = next(output_dir.glob("*-resume.pdf"), None)
    if target_pdf and target_pdf.is_file():
        return target_pdf

    raise SystemExit(f"Resume generation completed without a PDF in {output_dir}")


def run_coverletter_generation(repo_root: Path, record: dict[str, object], output_dir: Path) -> Path:
    script = repo_root / "scripts" / "generate-coverletter.py"
    if not script.is_file():
        raise SystemExit(f"Cover letter script not found: {script}")

    archetype = str(record.get("archetype") or "general").strip()
    role_num  = record.get("num")
    command = [
        _python_executable(),
        str(script),
        f"--archetype={archetype}",
        f"--job-role={str(record.get('role') or '').strip()}",
        f"--job-company={str(record.get('company') or '').strip()}",
        f"--job-team={str(record.get('team') or '').strip()}",
        f"--out={output_dir}",
    ]
    if role_num is not None:
        command.append(f"--role-id={role_num}")
    run_subprocess(command, cwd=repo_root)

    target_docx = next(output_dir.glob("*-coverletter.docx"), None)
    target_pdf  = next(output_dir.glob("*-coverletter.pdf"), None)

    if not target_docx or not target_docx.is_file():
        raise SystemExit(f"Cover letter generation completed without output in {output_dir}")

    return {
        "docx": target_docx,
        "pdf":  target_pdf if (target_pdf and target_pdf.is_file()) else None,
    }


def run_email_generation(repo_root: Path, record: dict[str, object], output_dir: Path) -> Path:
    script = repo_root / "scripts" / "generate-email.py"
    if not script.is_file():
        raise SystemExit(f"Email script not found: {script}")

    archetype = str(record.get("archetype") or "general").strip()
    minimal   = bool(record.get("minimal"))
    command = [
        _python_executable(),
        str(script),
        f"--archetype={archetype}",
        f"--job-role={str(record.get('role') or '').strip()}",
        f"--job-company={str(record.get('company') or '').strip()}",
        f"--out={output_dir}",
    ]
    if minimal:
        command.append("--minimal")
    run_subprocess(command, cwd=repo_root)

    target = next(output_dir.glob("*-email.txt"), None)
    if not target or not target.is_file():
        raise SystemExit(f"Email generation completed without output in {output_dir}")
    return target


def load_resume_metadata(output_dir: Path) -> dict[str, object] | None:
    path = output_dir / "resume-metadata.json"
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return None

def run_subprocess(command: list[str], cwd: Path) -> str:
    try:
        completed = subprocess.run(command, cwd=str(cwd), text=True, encoding="utf-8", capture_output=True, check=True)
    except subprocess.CalledProcessError as exc:
        message = exc.stderr.strip() or exc.stdout.strip() or "subprocess failed"
        raise SystemExit(message) from exc
    return completed.stdout


def _step(msg: str) -> None:
    print(f"  >>  {msg}", flush=True)


def _hr() -> str:
    return "  " + "-" * 50


def _row_get(row: sqlite3.Row, key: str, default: object = None) -> object:
    try:
        return row[key]
    except (IndexError, KeyError):
        return default


def _ascii_safe(text: str) -> str:
    """Replace Unicode punctuation that renders as garbage on Windows terminals."""
    return (text
            .replace('\u2013', '-')   # en dash
            .replace('\u2014', '-')   # em dash
            .replace('\u2018', "'")   # left single quote
            .replace('\u2019', "'")   # right single quote
            .replace('\u201c', '"')   # left double quote
            .replace('\u201d', '"'))  # right double quote


def print_existing_lookup(row: sqlite3.Row, heading: str, verbose: bool = False, repo_root: Path | None = None, show_jd: bool = False) -> None:
    num     = row["num"]
    company = row["company"] or "n/a"
    role    = row["role"] or "n/a"
    score   = format_role_score(row["score"])
    status  = row["status"] or "n/a"
    date    = row["date"] or "n/a"
    notes   = row["notes"] or ""

    stored_arch = (_row_get(row, "archetype") or None)
    archetype = stored_arch if stored_arch and stored_arch in ARCHETYPE_ORDER else extract_archetype_from_row(row)
    comp    = _ascii_safe(_row_get(row, "compensation_text") or "")
    wmodel  = _row_get(row, "work_model") or ""
    loc     = _row_get(row, "location_text") or ""
    jd_id   = _row_get(row, "jd_id") or ""

    print(_hr())
    print(f"  {heading}")
    print(f"  #{num}  {company}  |  {role}")
    if jd_id:
        print(f"  JD ID: {jd_id}")
    print(_hr())
    print(f"  {'Score':<12}{score}/10")
    print(f"  {'Status':<12}{status}")
    print(f"  {'Archetype':<12}{archetype}")
    print(f"  {'Date':<12}{date}")
    if comp:
        print(f"  {'Salary':<12}{comp}")
    if wmodel:
        print(f"  {'Work model':<12}{wmodel}")
    if loc:
        print(f"  {'Location':<12}{loc}")
    found_via = _row_get(row, "found_via") or ""
    apply_method = _row_get(row, "apply_method") or ""
    if found_via or apply_method:
        _src_parts = []
        if found_via:
            _src_parts.append(f"found: {found_via}")
        if apply_method:
            _src_parts.append(f"applied: {apply_method}")
        print(f"  {'Source':<12}{' | '.join(_src_parts)}")

    # Dimension scores if present
    dims = {
        "cv-match":  _row_get(row, "cv_match"),
        "role-fit":  _row_get(row, "role_fit"),
        "comp":      _row_get(row, "comp"),
        "work-pref": _row_get(row, "work_pref"),
    }
    if any(v is not None for v in dims.values()):
        parts = [f"{k}: {float(v):.1f}" for k, v in dims.items() if v is not None]
        penalty = _row_get(row, "red_flag_penalty") or 0
        if penalty:
            parts.append(f"penalty: -{penalty}")
        print(f"  {'Breakdown':<12}{' | '.join(parts)}")

    if notes:
        # Wrap long notes at 60 chars
        words, line_parts = notes.split(), []
        current = ""
        for word in words:
            if len(current) + len(word) + 1 > 60 and current:
                line_parts.append(current)
                current = word
            else:
                current = f"{current} {word}".strip()
        if current:
            line_parts.append(current)
        print(f"  {'Notes':<12}{line_parts[0]}")
        for part in line_parts[1:]:
            print(f"  {' ' * 12}{part}")

    print(_hr())

    if verbose and repo_root is not None:
        full = fetch_full_db_row(repo_root, row["num"])
        print_verbose_db(full, archetype=archetype)

    if show_jd:
        stored_jd = _row_get(row, "jd_text") or None
        if stored_jd is None and repo_root is not None:
            full = fetch_full_db_row(repo_root, row["num"])
            stored_jd = full["row"].get("jd_text") or None
        print("\n  JD TEXT")
        print(_hr())
        if stored_jd:
            print(stored_jd.strip())
        else:
            print("  (no JD text stored for this role)")
        print(_hr())


def print_existing_candidates(rows: list[sqlite3.Row], company: str) -> None:
    print(_hr())
    print(f'  Multiple matches for "{company}" -- narrow with --role or --id')
    print(_hr())
    for row in rows:
        score  = format_role_score(row["score"])
        status = row["status"] or "n/a"
        print(f"  #{row['num']:<5}{row['company']}  |  {row['role']}  |  {score}  |  {status}")
    print(_hr())


def fetch_full_db_row(repo_root: Path, role_id: int) -> dict[str, object]:
    db_path = repo_root / "data" / "job-log.db"
    with sqlite3.connect(db_path) as conn:
        conn.row_factory = sqlite3.Row
        row = conn.execute("SELECT * FROM roles WHERE num = ?", (role_id,)).fetchone()
        reqs = conn.execute(
            "SELECT requirement_name, kind, priority, matched_normalized FROM role_requirements WHERE role_id = ? ORDER BY priority, requirement_name",
            (role_id,),
        ).fetchall()
        my_skills = {r[0]: r[1] for r in conn.execute("SELECT skill_normalized, level FROM skills_mine").fetchall()}
        my_capabilities = {r[0]: r[1] for r in conn.execute("SELECT capability_normalized, level FROM capabilities_mine").fetchall()}
        my_qualifications = {r[0]: r[1] for r in conn.execute("SELECT qualification_normalized, met FROM qualifications_mine").fetchall()}
    return {
        "row": dict(row) if row else {},
        "requirements": [dict(r) for r in reqs],
        "my_skills": my_skills,
        "my_capabilities": my_capabilities,
        "my_qualifications": my_qualifications,
    }


def print_verbose_db(full: dict, archetype: str | None = None) -> None:
    row               = full["row"]
    reqs              = full["requirements"]
    my_skills         = full.get("my_skills", {})
    my_capabilities   = full.get("my_capabilities", {})
    my_qualifications = full.get("my_qualifications", {})
    LEVEL_NUM = {"exposure": 1, "basic": 2, "intermediate": 3, "advanced": 4, "expert": 5}
    MET_LABEL = {"yes": "yes", "partial": "~", "no": "no"}

    print("\n  DB ROW")
    print(_hr())
    skip = {"num", "first_seen_date", "last_updated_date", "jd_text"}
    if archetype:
        print(f"  {'archetype':<24}{archetype}")
    for k, v in row.items():
        if k in skip or v is None or v == "":
            continue
        print(f"  {k:<24}{v}")
    print(_hr())

    if reqs:
        print(f"\n  REQUIREMENTS  ({len(reqs)} extracted)")
        print(_hr())
        by_kind: dict[str, list] = {}
        for r in reqs:
            by_kind.setdefault(r["kind"] or "other", []).append(r)
        for kind in ("skill", "capability", "qualification", "unknown", "other"):
            items = by_kind.get(kind)
            if not items:
                continue
            print(f"  [{kind}]")
            for r in items:
                matched = r["matched_normalized"]
                if kind == "qualification":
                    met = my_qualifications.get(matched) if matched else None
                    badge = MET_LABEL.get(met, "?") if met else "?"
                    match_label = f"  -> {matched} [{badge}]" if matched else "  -> (no match)"
                    print(f"       {r['priority'] or '-':<10}{r['requirement_name']}{match_label}")
                else:
                    level_map = my_skills if kind == "skill" else my_capabilities
                    level = level_map.get(matched) if matched else None
                    num   = LEVEL_NUM.get(level, 0)
                    match_label = f"  -> {matched}" if matched else "  -> (no match)"
                    print(f"    {num}  {r['priority'] or '-':<10}{r['requirement_name']}{match_label}")
        print(_hr())


def print_summary(metadata: dict[str, object], dry_run: bool, verbose: bool = False) -> None:
    db      = metadata["db"]
    role_id = db["role_id"]
    score   = format_role_score(db["role_score"]) if db["role_score"] is not None else None
    status  = db["role_status"] or None
    ai      = metadata["ai_provider"] or "n/a"
    label   = "DRY RUN" if dry_run else ("Logged" if db["status"] == "logged" else "Intake")

    print(_hr())
    id_str = f"#{role_id}  " if role_id is not None else ""
    print(f"  {label}  --  {id_str}{metadata['company']}  |  {metadata['role']}")
    print(_hr())

    rows_out: list[tuple[str, str]] = []
    if score is not None:
        rows_out.append(("Score",     f"{score}/10"))
    if status:
        rows_out.append(("Status",    status))
    if metadata.get("date"):
        rows_out.append(("Date",      metadata["date"]))
    rows_out.append(("Archetype",     metadata["archetype"]))
    rows_out.append(("AI",            ai))
    if metadata.get("compensation"):
        rows_out.append(("Salary",    _ascii_safe(metadata["compensation"])))
    if metadata.get("work_model"):
        rows_out.append(("Work model", metadata["work_model"]))
    if metadata.get("jd_id"):
        rows_out.append(("JD ID",     metadata["jd_id"]))
    rows_out.append(("Output",        metadata["output_dir"]))

    for key, val in rows_out:
        print(f"  {key:<14}{val}")

    print(_hr())

    if dry_run:
        print("  DB            (dry-run)")
    else:
        db_val = "written" if db["status"] == "logged" else db["status"]
        print(f"  {'DB':<14}{db_val}")

        if verbose:
            resume = metadata["resume"]
            cl     = metadata["coverletter"]
            em     = metadata.get("email") or {}

            def _artifact_path(art: dict) -> str | None:
                return art.get("pdf") or art.get("docx") or art.get("path") or None

            for label, art in [("Resume", resume), ("Cover letter", cl), ("Email", em)]:
                path = _artifact_path(art)
                if path:
                    print(f"    {label + ':':<15}{path}")

    print(_hr())

    if verbose and db.get("full_row"):
        print_verbose_db(db["full_row"], archetype=metadata.get("archetype"))


def format_role_score(score: object) -> str:
    if score is None or score == "":
        return "n/a"
    numeric = float(score)
    if numeric.is_integer():
        return str(int(numeric))
    return f"{numeric:.1f}"


def open_file(path: Path) -> None:
    import os
    try:
        if sys.platform == "win32":
            os.startfile(path)
        elif sys.platform == "darwin":
            subprocess.run(["open", str(path)], check=False)
        else:
            subprocess.run(["xdg-open", str(path)], check=False)
    except Exception:
        pass  # non-fatal — user can open manually


def cleaned_lines(text: str) -> list[str]:
    return [line.strip() for line in str(text).splitlines() if line.strip()]


def clean_label(value: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"^[^A-Za-z0-9]+|[^A-Za-z0-9]+$", "", str(value))).strip()


def normalize_text(value: str) -> str:
    return normalize_search_text(value)


def slugify(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", str(value).lower()).strip("-")
    return slug or "unknown"


def date_stamp() -> str:
    return datetime.now().strftime("%Y-%m-%d")


def today_iso() -> str:
    return date_stamp()


def html_to_text(html: str) -> str:
    text = re.sub(r"(?is)<(script|style).*?>.*?</\1>", " ", html)
    text = re.sub(r"(?i)<br\s*/?>", "\n", text)
    text = re.sub(r"(?i)</p>|</div>|</li>|</section>|</article>|</h[1-6]>", "\n", text)
    text = re.sub(r"(?s)<[^>]+>", " ", text)
    text = unescape(text)
    text = re.sub(r"\r\n?", "\n", text)
    text = re.sub(r"[ \t]+", " ", text)
    text = re.sub(r"\n{3,}", "\n\n", text)
    return text


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except AmbiguousExistingSelection as exc:
        company = "requested company"
        if "--company" in sys.argv:
            try:
                company = sys.argv[sys.argv.index("--company") + 1]
            except (ValueError, IndexError):
                company = "requested company"
        print_existing_candidates(exc.rows, company)
        raise SystemExit(0)
    except KeyboardInterrupt:
        raise SystemExit(130)
